from datetime import datetime
from datetime import timedelta
from io import BytesIO

from django.conf import settings
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMessage
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.cache import never_cache
from pytz import timezone
from xhtml2pdf import pisa

# import win32com.client
from acc_eodc import settings
from .forms import ImporterForm
from .models import Importer
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter



def send_excel(request):
    return render(request, 'excel.html')


def create_excel_file(request):
    start_day = request.POST['start_date']
    last_day = request.POST['end_date']
    bond_type = request.POST['bond_type']
    print(start_day)
    print(last_day)
    data = Importer.objects.filter(lic_date__gte=start_day, lic_date__lte=last_day, radio_choice=bond_type)
    print(data)
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Add headers
    ws.append(['IMPORTER NAME', 'IEC CODE', 'ADDRESS',
               'LICENSE NUMBER', 'LIC DATE', 'BOND NUMBER',
               'BOND DATE', 'BOND AMT', 'DUTY SAVED',
               'ITEMS IMPORTED', 'EMAIL ID', 'BOND TYPE',
               'EODC PRODUCED', 'DGFT ACK PRODUCED', 'LETTER ISSUED',
               'SCN ISSUED', 'PH1 ISSUED', 'PH2 ISSUED', 'OIO ISSUED'])
    first_row = ws[1]
    for i in first_row:
        i.font = Font(bold =True)
    column_widths = [17, 14, 40, 17, 14, 14, 14, 14, 14, 17, 40, 14, 17, 20, 14, 14, 14, 14, 14]

    # Loop through the columns and set their widths
    for i, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width


    # Add data from the database
    for item in data:
        ws.append([item.importer_name, item.iec_code, item.address,
                   item.license_number, item.lic_date, item.bond_number,
                   item.bond_date, item.bond_amt_executed, item.duty_saved,
                   item.items_imported, item.gmail_id, item.radio_choice,
                   item.is_eodc_produced, item.is_dgft_ack_produced, item.is_letter_issued,
                   item.is_scn_issued, item.is_ph1_issued, item.is_ph2_issued,
                   item.is_oio_issued
                   ])
    subject = 'data between ' + str(start_day) + ' and ' + str(last_day)
    message = f"please find the data between " + str(start_day) + " and " + str(last_day)
    from_email = settings.EMAIL_HOST_USER
    recipient_list = [settings.office_mail_id]
    email = EmailMessage(subject, message, from_email, recipient_list)

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    email.attach(str(start_day) + 'to' + str(last_day) + '.xlsx', buffer.read(),
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    email.send()
    return redirect('home')


import pypandoc
import tempfile
import os


def generate_docx_from_html(html_content):
    # Create a temporary output file
    output_file = tempfile.NamedTemporaryFile(suffix='.docx', delete=False)
    output_path = output_file.name
    output_file.close()

    # Convert HTML to DOCX and specify the output file
    pypandoc.convert_text(html_content, 'docx', format='html', outputfile=output_path)

    # Read the generated DOCX file as bytes
    with open(output_path, 'rb') as docx_file:
        docx_byte_data = docx_file.read()

    # Delete the temporary output file
    os.remove(output_path)

    return docx_byte_data


def generate_pdf(html_content):
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("ISO-8859-1")), result)

    if not pdf.err:
        return result
    return None


@never_cache
@login_required
def send_reminder_emails(request):
    # Get today's date0
    today = datetime.now(timezone("Asia/Kolkata"))

    # Calculate the date  days ago
    epcg_time_gap = today - timedelta(days=settings.epcg_time_gap)
    epcg_letter_time_gap = today - timedelta(days=settings.epcg_time_gap) - timedelta(days=settings.letter_time_gap)
    epcg_ph1_time_gap = today - timedelta(days=settings.epcg_time_gap) + timedelta(days=settings.ph1_time_gap)
    epcg_ph2_time_gap = epcg_ph1_time_gap + timedelta(days=settings.ph2_time_gap)
    epcg_oio_time_gap = epcg_ph2_time_gap + timedelta(days=settings.oio_time_gap)

    decc_time_gap = today - timedelta(days=settings.decc_time_gap)
    decc_letter_time_gap = today - timedelta(days=settings.decc_time_gap) - timedelta(days=settings.letter_time_gap)
    decc_ph1_time_gap = today - timedelta(days=settings.decc_time_gap) + timedelta(days=settings.ph1_time_gap)
    decc_ph2_time_gap = decc_ph1_time_gap + timedelta(days=settings.ph2_time_gap)
    decc_oio_time_gap = decc_ph2_time_gap + timedelta(days=settings.oio_time_gap)

    # for epcg
    letter_records = Importer.objects.filter(radio_choice='EPCG', lic_date__lte=epcg_letter_time_gap, is_paused=False,
                                             is_closed=False,
                                             is_eodc_produced=False, is_dgft_ack_produced=False, is_letter_issued=False)

    scn_records = Importer.objects.filter(radio_choice='EPCG', lic_date__lte=epcg_time_gap, is_paused=False,
                                          is_closed=False,
                                          is_letter_issued=True, is_eodc_produced=False, is_dgft_ack_produced=False,
                                          is_scn_issued=False)

    ph1_records = Importer.objects.filter(radio_choice='EPCG', lic_date__lte=epcg_ph1_time_gap, is_paused=False,
                                          is_closed=False,
                                          is_scn_issued=True, is_letter_issued=True, is_eodc_produced=False,
                                          is_dgft_ack_produced=False, is_ph1_issued=False)

    ph2_records = Importer.objects.filter(radio_choice='EPCG', lic_date__lte=epcg_ph2_time_gap, is_paused=False,
                                          is_closed=False,
                                          is_ph1_issued=True, is_scn_issued=True, is_letter_issued=True,
                                          is_eodc_produced=False,
                                          is_dgft_ack_produced=False, is_ph2_issued=False)

    oio_records = Importer.objects.filter(radio_choice='EPCG', lic_date__lte=epcg_oio_time_gap, is_paused=False,
                                          is_closed=False,
                                          is_ph2_issued=True, is_ph1_issued=True, is_scn_issued=True,
                                          is_letter_issued=True,
                                          is_eodc_produced=False, is_dgft_ack_produced=False, is_oio_issued=False)

    # for DECC
    decc_letter_records = Importer.objects.filter(radio_choice='DECC', lic_date__lte=decc_letter_time_gap,
                                                  is_paused=False, is_closed=False,
                                                  is_eodc_produced=False, is_dgft_ack_produced=False,
                                                  is_letter_issued=False)

    decc_scn_records = Importer.objects.filter(radio_choice='DECC', lic_date__lte=decc_time_gap, is_paused=False,
                                               is_closed=False,
                                               is_letter_issued=True, is_eodc_produced=False,
                                               is_dgft_ack_produced=False,
                                               is_scn_issued=False)

    decc_ph1_records = Importer.objects.filter(radio_choice='DECC', lic_date__lte=decc_ph1_time_gap, is_paused=False,
                                               is_closed=False,
                                               is_scn_issued=True, is_letter_issued=True, is_eodc_produced=False,
                                               is_dgft_ack_produced=False, is_ph1_issued=False)

    decc_ph2_records = Importer.objects.filter(radio_choice='DECC', lic_date__lte=decc_ph2_time_gap, is_paused=False,
                                               is_closed=False,
                                               is_ph1_issued=True, is_scn_issued=True, is_letter_issued=True,
                                               is_eodc_produced=False, is_dgft_ack_produced=False, is_ph2_issued=False)

    decc_oio_records = Importer.objects.filter(radio_choice='DECC', lic_date__lte=decc_oio_time_gap, is_paused=False,
                                               is_closed=False,
                                               is_ph2_issued=True, is_ph1_issued=True, is_scn_issued=True,
                                               is_letter_issued=True, is_eodc_produced=False,
                                               is_dgft_ack_produced=False,
                                               is_oio_issued=False)

    base64_image = "data:image/jpeg;base64,iVBORw0KGgoAAAANSUhEUgAAAzgAAAEECAYAAADkq5F9AAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAAEnQAABJ0Ad5mH3gAAL9ESURBVHhe7J0HgB5Vuf6fr2/fTXbTN5teIIUQkgAhhE4AxYKggFflWlDsV1G8XryWy1VRrH8VxXKxgShiA+mhBAIkEAikkt42ye5me/3a/N9ndg+eDPOVTSHJ5v3Bybczc+a0OXPO+8w5cybgCFAURVEURVEURRkABPt+FUVRFEVRFEVRjnlU4CiKoiiKoiiKMmBQgaMoiqIoiqIoyoBBBY6iKIqiKIqiKAMGFTiKoiiKoiiKogwYVOAoiqIoiqIoijJgUIGjKIqiKIqiKMqAQQWOoiiKoiiKoigDBhU4iqIoiqIoiqIMGFTgKIqiKIqiKIoyYFCBoyiKoiiKoijKgEEFjqIoiqIoiqIoAwYVOIqiKIqiKIqiDBhU4CiKoiiKoiiKMmBQgaMoiqIoiqIoyoBBBY6iKIqiKIqiKAMGFTiKoiiKoiiKogwYVOAoiqIoiqIoijJgUIGjKIqiKIqiKMqAQQWOoiiKoiiKoigDBhU4iqIoiqIoiqIMGFTgKIqiKIqiKIoyYFCBoyiKoiiKoijKgEEFjqIoiqIoiqIoAwYVOIqiKIqiKIqiDBhU4CiKoiiKoiiKMmBQgaMoiqLAQRJJp9t1iVQPkukEEskkUmkHqSSwedMW/OPePyCVakIi0Sz7upDmAUf+F5dOp12nKIqiKEcaFTiKoiiKEEAwEOx1wQACAXGy1xHREgg6uP+B+7F61cuor6tFJBJG2knJUQoaOlE4iqIoinKUoAJHURRFccVMr6SRjkHEDUWL44h4ESGzZ/du/PWeP6MgFsaO7ZuQTHQhFBThI+KGfhwO4SiKoijKUYIKHEVRFKVvDCYgnUJvtxCQPZFQEOFIGI8/+jBWvfQSSouj6OhoRktzPYKhMAIBB4GgnOM6Oc8VRoqiKIpyZFGBoyiKorijME0tzWhtb0NnRwd6OjrRJtu7t23F3X+4E+l4D4LpFAoiEWxYvw4Ne3dhX0M96uv2oK5uL+rr69Hd090XmqIoiqIcOQLSqencAkVRlOMcLiqwdOlTePSRR1BaVIyxNWMwqKwMTiKJ3/32t3j4oQfxre99CSNHDcVPb/0V6hpaESsqR1lFJUaPGYtJk6fg4osvwZjRY/pCVBRFUZQjgwocRVEUBWn5Lx7vwqvr1uNXv/gFXn7xRYwYMhSzZ85EVWUlHnrgfpx2xklYs24ttmzdhbHjJ2LilOkYVFmFcZMm4qRZs1BRVoFQKCKh6VQ1RVEU5cihAkdRFEVBIhkHX6MJijbp6ezCw/c/gF/d9jNs27wZc2efgjE11WjvbEFXTzdmn3IqqmvGIVJQjJoJ4zBs5AgEQgGkpTspiBRJaCpwFEVRlCOHChxFURQF6XQKqVTCXVwgHAwBoQg2vvg8lj/7DEaPGoXbfvpTzJ8/H888twxLlz2Pc86/ELf84IcoKy9FV7Ib4UgQSelOCiMlEpoKHEVRFOXIoYsMKIqiHPPwOZXtvHAfv1tjH+My0En5NfuDCAQiCIZCSDlxpJNdaO3qxOgx4/GnP9+DxU8sQUXlELznfe/HsGGjsHjxEuzd0wAnEEI0WiSaJoJQMCrh9Iqb3mdnxu0PD9nOz4+iKIqiHCgqcBRFUY5xer9F43UUMH1OREwaPe5+813OdDohPwkkU91IJHgsIPsobuR4MI229ka0dLTjr/98AHf/40GcMOsUREtKUVRWio98/OMoLCnHAw88ikA4JIGF4aRjCDhRpFJpCac3DRwVojPpSafFme+DviZs7PTKHv1uqKIoinKQqMBRFEUZcPQKh/2dETvyl/zD92XSTgqhUBihYBjJZMr9jA2/hBMKFmLfvk48/PCTeOHFNfjoxz+D2++4C29/57swoqYG48dPxFVXXYVVq1ahtbEDAelJQgEHyURCQnejcsNLJjlC1LsjnU5KvKJuApIW9jzuQI945Dadm0ZFURRFOXhU4CiKohzzsCl/TTX0YYsHbvY29+4oCQ+J/wDCSCUDImzCiMVC4GBMIB1Ce2sKD/xzKTZtrMOHrv0kLrviCjTua8PLK9eisyOJotIKnHDCdAktiJdfXo3uzjiclIPCgqgIpqCEHxRBI6GHYwiKeAqIAur9GKgjf9OlxVFQiesTX5D9vb/yY2dDURRFUfqJLjKgKIpyjLN/M27+plgwf/dOAYMThSPCIyRCpldEpNHdlRQhsxvr1q3Fhg1rsWnDTuzY3oCVL7+M4qIid+noppYGEUIpBJ0kgiJWYrECxKIxdxRo2PAhGDuuBjWjR2DSlIkYN3Eipk07EaOqhyMSZUQCkyFp4QhO2km7Yqd3Jx0FDrdFgDn0z3eBVOEoiqIoB44KHEVRlGOc3lac08B6xYI71axPJLhNvIiLYOBfYmP7tj1Y8eIKPPbY41j+3EoROHvR1dUuB1MoKqhCQWwQKiurUFYeQ1FJGAUFURQWlCCdCiCd4Hs0abS2tiDe04nGpr1oa21Ai7hEKgFEgnJeBWbOmInTTj8Np4ubOfNEDB1e5kbvvp8j8fC9Gy5LzQT1To0LuaM+/CvIVdwURVEU5QBRgaMoinKMw3ddgqIWepvzXpGQSvaKnXCY+4HGhhYsX74Kf/vrfXj66aexe88udHfHMXzIWIyuPhETJ07AyOoRKCsdhXCwEOFQSgRLA3oSrejuSiDRzfd3SkShhBGNRlBUGEFhLATRPojHm9He0YjG5nps2b0TmzZtwa5de9DS3IqS0hJMmDAOC8+ajzdfegFOmjUN5eWF7mIEqXRSxE3viA7zEInE3DRzSpuO4iiKoigHigocRVGUYxy+wM+pXW5rztXQUhBhI6JBfmt31eHPd9+Lv9xzH5579gUEQwUYW1ODOXNPxkkzTkFF+Sj0dMawZcdObN+1G3t2t2BPbS06Onegq2s7FpwxG0iF8dQTK5EOlEs0IkJcYRJHQcRBeUkEI0aU4eKLz0ZpWRGauroxfGQ1kokQtmzejhUrXsQrr6zE3qZdKC8uwoIzT8e7rnwHLrr4LFQNLUYinkYg6CAUosjhyI4jQieiAkdRFEU5YFTgKIqiHOMkk3F31IMz1MLBsLtv29YG/OHOv+BX//c7bN+2C5WDhmH+6WdhzimnoapqMLZt34J16zZjw7o92LurCz1OFJ1JB8l4QIRLGqXlzSgrb8WwymKcPX8RdmzpxPOrNiNaVIp0ohtd7U0IoAvtzbWYNHk4zlowBxs2b0QiVIxnl63EiOFjMHP6KZg6dRpKSoqxbv1qPPPMEqxa/ZIIr05Mnz4Z1173Hrz9HRehsqpMBBFXWBP95C5G0LsggqIoiqIcCCpwFEVRjmk4dSwtAqF3ildjQxv+9Md/4Fe//D1WvrgOI0bU4OyzzsXsk08TPwG88PxLeGnli6ivr8PMGaegpzOK1S/vRCoyCD2IYPjQEZh/6okoKd2LwYMbMWJIKbas3SNCRkTKzgYMGjYEwXRcRE4HakZVojAax+RJw1C3ZzPWb9iIiTPOxK9+dSeaGrukgylAeXkVxowZg/kLTsP4cTWob6jF4sfux8pXnkd3T7Psn40Pf+R9uOjic1FcUoCe7gRiBZG+vCmKoihK/wl9Rej7W1EURTkK6X0M5c4/E8d3a0jAfY+Fyz1zCehQOIxnl76MT37yP/HjH/0abS1JvO3Sd+O97/4wyoqH4dFHnsbdf7ofa1ZvRmtrm4iJiPtuzPhx07Bndw+aOgJArBSdHW0oKnQwamQUL614HHPnzEZ9XScSyVI8t2I1Zs2ejTIRIkWFYRFD5eK/EYVFMTy7/DnMOWUBCqIjkOoOYvbMU5BMBrCvpQdbdjRi2Yp1WLVmmwieSixadAFOOnkWmppa8Nyzz+OBBx7D9q17MX3aCRgyrELO6/1eD//pfQbHhQh0ypqiKIqSHypwFEVRjmJo33PFsV5hwxXIuOQzkEhwyWYx+sXFu1O45ds/x+c//1WsXb0dp5+2CB+97gsYMXQS7v37E/jLPfdj0+atSCULMGLkGMybNxuzTpmCqmGDMG7siXjxhW1o7AwjGSxCWsIdM2oIZpwwVgRUHOPHT8NTz6zH5Gln4tkX1ojwGY36vXVo2FuLmppRWLd2Hbok/meXvYxpM+fjgX88hb276rFwwQI5Pg4nnnw6omXVqK3rRnNzCqteWYvVq15BSWkprrj8SoyuHo/t2+qw9KlleOihxRg0uBxTT5iEUDCIpLsQQQpwv5/DaWv/EjkqdxRFUZRMqMBRFEU5qukdwXCdu9wz37VxEA5F3HdVduzYg098/Av42U9/haJYJd7znutwwfmXYenSV3DHnX/Hq6/uRDAcFrExFO+8/J2YOuVErHjpOdSMH4m169ahsLAKsdhQrH21CYHgYDjpCDrae9DTkRYRFUZbRwR76hys39iC2j3NWLN2A+r2NKCiolLCHI9T5sxHW1sKzS1JTJg0Q8TQs3jP+96NpqZ9ePyp5zD7tHOw8pVt2LunHcF0IcIIob2lDWtXrZG0rcMpp8zFokWXoL2jHc+vWI5HHn3IHWGaM/cUlBTH3NXVuGy0q28MTu9S2IqiKIrih76DoyiKchTD0Rt3BCfAEZyA+6FORwz8cDiE55etweeu/y88teRZzDhxLt77bx9FMl2GO35/H17dUCvCphjRgmK86ZJzcNIJQ7D40Sdx+mkL8bs//ApnnnMKCouKEYsOR2nRBHzre39Dd0L8ByNIJ9sQTLWIpGpGUWkUnXF+oJOLFwQRCgbETxzVoyokLW0oLylBSNLS0tqKkdXDsXT5I3j7my8SMeNg7ITpWLe5FXf9+VFEIxWIOBFEUp2YNnWo+N+FbTtXo6A4hIULz8Clb7sIS599GH/+62/Q2LgL73rXZfjGN76MmjFDJG5mnfnvE3qS/9e+66MoiqIoHlTgKIqiHMXw6//8MCbFDd+1cfi+jYiMhx5Yik9/+nPYsmUnzjrjfFx91fvx4gsb8ec/LUZDExApEGEQKEA6VICa6kG48tK5ePbZ5SgoKMOrm9dhd902TDnxROzZ24VYcDT2NCTQ3plEQaRI4kkC6S4RFj1IBboRjAQQT6cRFvEUcJPSJQKmC+FIUkROCpFw1F2SOiDpSgSbgVQ3Tpw8BYsueAt+fcf92LMvgaCIm5iIkoqCFE49pRpDR5Rg4+aVeGnVC+iO92DCxHH49w9chaaWWtx5x23YuHE1zlw4Hz+59RaccGINUq6+k3S5CQBC4EIEutqaoiiK8npU4CiKohzFcPQmLSKndzEB+TcUxMMPPI0PXfspNNZ34OJFl+Idl71LhM0/8OjDzyKdKkdV1RQ0dQbQ1h1COliARE8rTp5ShQVnnCGCphGFxQXYWbsNS555BrHoYCTiBSKiouCgSED+Syc5FS6AlHQPTrBHwugWsRRH2BkMJCLio0dSExeRkUIoFEIgEJakRRBPpiE6BiFHBE0qgdKSwWjtBnpSMYQ5tS7ejrdfdKrEWYeXVy3F+YsWMEL89a/3Ym/dPlQNHYz3vu9dqBwcws9+/n0ROetw8pwZuP32H2HqCaNF5EhZSLxSEIjwezwqcBRFURQf9B0cRVGUoxrHfQ+FYxYBUSBPPL4MH/vY51C/uxVvvuQKvOniy3DHb/+IxxcvQTRcgBNPnIFTT52PDZu3o7NH1EOwBJFICfaJgFizbht272mW8MIiJkZh65Y6OKmIiJOgK2hE0rgjRkFHtvmOj4iWSDiCcCSERKJLREVQnIgaETjhQAJhzhILppFMJyWVHF2KiAgpRiggLh1DTxckxJiIJBFPIk0igTRmnTQere2bcML0sRgyrBw7d25DUaH4D0awd/derHp5FcaNq8GFiy7A5s3bsGb1Gqx8eTXmz+f3e8ol7b0jOJyipiurKYqiKH6owFEURTnK4bS0iIiMtWu24EMf/CQ2vrpTxM27XIHz29/8CU8vfRYlIhIuXnQhTjt1NrZs2YS5c+ejszuNhn1dcn4MoVAMSSeE9o4eETn12LBxiwiUGBKphAiSFMIiUEQyyH9xhEXEQPaHgyE46QSKC0OYPH6k/KaQ6NmHkcNLRJSk0ZNockdrIuGACJx0r+AIFCLNMEXaBIJhyhoROCJv0nERU90YN3YIQpEEho0cgtWrXxLB9gQqB1fh9FNPR7y7CztqN+PV9WsxenQNLjj/Qqxb+6oInFewdet2nHvuWSgrK0ZayiMcorhRgaMoiqK8HhU4iqIoRzMORUMQjQ2N+MhHPoNnn3kBZy+8BO++8oP43W/uxvPPrxLxU4izzlqAsWOHobw0jOHDK7Fp01bMmjUfL6zYIOKjBAkRBWkJKihiJBikNOCISxKRaEKER7tsd4tQSWDKlCE4Zd4E1IwehN27NomvTrzpovkYUimiItyGwRUhXPGORRg6pABd7Q04ZfZ0JJLdiEYdVwwl0j0ieroQC4sASoq4ClDqcJlniGByUL9nL0aNHo/6hgasXrMaV1z+DoypqcGu7dswa8YUFMYC2LR1I1atXoupU0+UfJ2NF1e8iFWrVqGjoxNnnb1AxI2IMVfgKIqiKMrrUYGjKIpyFMMF1NKpFP7na7fgjt//CdNOOAUf/PeP4+EHn8QTTywHZ6/Nnz8fo0YOEQFQJ/sewGnz56Gr28HzKzZib103Uk4M6VAIaREfjiMCxP2uDL+jExdxFBeBMVjOLxVRU4QTpg9Cd3I7ysqSKC5Oo7FxC+bMrsHWLc8hEd+HqspC8VuObVtWo721Hm972yUoLoqJqKpCd087Fp57GtLJVjQ17HIXQ+iRtAfDEYk3LQIqJH5S2LixVsLbhqHDKjFh/Hj86a47cdYZp2HYsHJUlEfR0tqGXXv2YdPGzZgzZx4mT5mMlS+9hBUvvoDq0dWYM/fEvtJRFEVRlNejAkdRFOUI8q91XvjryLa7TBlS/MilOI5W/ONvD+OrX/5flBZX4mPXfQH1e7vx69/+CYm0g3mnzcaIIRWor9+FM8+Zj1BRKXbI8ZdW7cXqdXtFUHDiWRABCTcUSCMYTIm46URhUQLRaAsKgm14//vegkkTyzFhXAn27F6NkSMLEXRa0NS0CRs2PoP2tu0YXV2CQHIPmvauRjjQiA7ZV17qoLF+kwigKVj+3MOS3mbMmDUJu3Ztwrhx1SgtK0RLcx1CoThXtEGCrxIh2juq44QQjpZg9+5G1O7ag9knz8AzSx+RvzfhiiuvwvqNGyUte7Ft2y684+2Xo6enB2vXrsFLK1/EOeechaphg/nJU8QTid4yC3AhBu5h+XFZbRFxkm9FURTl+EMFjqIoylGBETj85Ycs+Z5JCDu3N+CTn/wcdu1swBWXX4OxNTPwwx/+Et3dCUyaOgFz5s1A895dmDVrGpa/uBLPrViHpcs2Y+uOHvTEiyQcERTpJAqjve/XpFLtQLAD8+ZOxLuvvBCdrVvQ0bINFaXdaG/fhE0bV2DH1lVYs/oZ7Ny5GmNqylFUlEYy2YLCcBui4XbUN2xFXd1GrFmzHOteXYm9ezZh29a1mDPnBDzzwtOIREN421veguFDByMWAzZvfhWhoKTD4WptYfedHAQi6OpMYN8+fm9HxIjTg3df9Q6UlxVhyTNPYdbsU7B7zx5s37YTnZ3duPyKd+LVV9e67xfVN+7DhRefJ8Ip6H7sNC35C4U5Za2v/Bi8u2iCfitHURTleESXiVYURTmC/KsJ5odeuC3Wuev4fwDf/MZP3elpJ0yagc999sv47W/+gueefVFERyHOPPt0rFu3Ele8+a1Yu24NWkX0rFq/G3vqaOAPRypdgmSK65slEYvEkUh1ifhwRBw14OSZ1SJwzse+urXYvGEFVq96CvuatqCmZhhOOGESTpo1A6Ora1BaWoaysgpJTBDRKIVOD+KJOJpbWlFf34TNm3Zg2XMvyu92xJNhtCYiuPTNV6MwNgSDK6rxz/ufRFOzyJroSOyVdLV1hhEIxkSQcOW2lIivDoSDLXjThXNRFGvDK68swayTTsC06TPwwvMv48GHH0MkEsNHrvuQlFAXvv//vgmEuvH/fvoNXPmuS1lMUnQicEKcdsdC456Au8qaeORRRVEU5ThDBY6iKMoRpLcF5j+9AiclgsRxQu7KZKtXbcKbLnkvGuq68R+f+oIIgMH43nd/5L6Tc8lF52P4iEG4776/YtzosRg/cTI64xE8sXQd9uwNiLApFxdCMJxCMNSJwWVJnLngVDQ11uHll5eKAOnAmy6ajYce+A16Ovdi5qxqnHvhHBEXMzGqeqSImShKiksRDheIKIpJOAVIpuMiTEJIOyI1OjqRSiTR05VAc1MrNqzfhMeeWIrHn34BjfviOGHybIwYPlmETSWmnrAAtXuCuP23D0kYg8UV9ubbSSEa4ehNK4oLuzBlQiXOPfsUbNv6Ala+uAyTJs3Evn3tWPHSy5gwcTw+ff3H8Iv/uxVPPPUAFpw1A3++5/9QXl4qckbyGEyLqGEJ9gqc3tEbnaKmKIpyPKJT1BRFUY4GAnxvRGROqteFw0F895af474HFmPOrDNx6aVX4rbbfivipQ6jRg3DxEmj0Nm+D2ctOE1MegcrXl6N9a/udEdLEolCpJyIGP0JESfdiMd3YeL4Ypxz9skojHZh9kk1WLf6Cby44n4RISF85MNvx7uvuRSz5s1AzbjRqKgchMFVVSgur0BBRQUiRSUIFxQgWliMSKwY0YIYCosKUSCupLwEpSVFbppmzpyKU+dNRzLRiWefeRKdnY0oLS3AuPHjsWz5SsleAUaPmoi9e5sR4giLw+WlJduiQ3oSccSihVi1ai1Ki0I4/9xz0Nbag8mTZ2Dtus2S7wYMqqzEqaefLmEtFxH0KqZMmYrZJ08RIZdy31XqFTcsRv5N5yoeRVEU5ThDBY6iKMoRRwxzTq/iX2L0B8Xi37ZlN/7rC99EOhnAVVe9F9u378PDjzztjqhc+pZFYtyPxY7tm7H40UcRKSzF6QvOxZy5Z6K4dBhWrd0kIaUQjcUlvHpc9rbTMWRwAq+89CQqyjj6sh3r1jyB88+bhc9+9t9xytwpGDG2GkWDhqFMXChWJsJoEAKhIuxrbMfTTy/Hiy+uEmHxPNasXSvpi2DosFHuqE4oUuBOISsqLkUpxU5xENOnTcQEEUrPPfcMXl71EoqLi1FWMQyTJpyIdes2oaOD3+ZJi7AJIplMgh8wddIBNDY1SzmEcembLsCTTzwuwmi6uBNF+JTilfXrsG9fExYtWoS9e/Zg85ZX0drShMsuf4scD7ujSr0jOILTK270Q6CKoijHJ+wFFEVRlKOBvhnD4UgA9/7jcWzbXouxY8di4qSJWPLU80imCjBhyiyMnzoD9c2dmHfaeZg2fQHqGhN49Inl2LqjDt09IhiCcRTEUgihCSdNH4bB5Qmcd/YsnDxzJNatfhT33vtzXHHFQnzm+vdj4tQxqJk0GeWVoxAtHIaEU4a2jghe3dKAnlQpHnhkOb729f+HJ556Ce2dwF/ueQg//dnvkJK0tLemUL+HixYUS6KL3VXRhldXY8yEsTjvwgX46v98BlOnDsFjj92NqsEOXnn5cdTueEW0XIOkrQVOqg2hYBrJBFc+i0q6S9HWGUJC4j157oVYt34vHnrkOaScIgwqG4kdO/Zi1SvrsHDhuSguGoQXXliFZ5audJej/peYMSM3Km6ORdzFIfqcoijKgaICx+Kiiy56rWH9wx/+0LdXURTlcENhw1GcAMLhEDraEnjowcfFcI9hzinzULd3HzZv2Sst9mCEIlVY8swaLH9hM7bt6MGESWfBCYxEY0sEe+o7seTpJQiHwkgn2hF0OjH35MmYdeIYxDvqsLd2DdatexbvetdCvO+aSzB0RCmGjhohwqZCwi5HKFSCtpYkvv6N7+NnP/0Nbvn2j/DKyldxwflvQkFBGYoLBmHK5BkoLRmCP9z5N9x6669x45e+jkceeRppFCIQLkIgUozywUMxcswYzJo9DZ+74aMYNrwQ/7z3t0j27EUY+zCsIojT5kxG1SBOo+MHRgPutDxOYYtGKrDi5W1YvnI7nlmxUYTbCvz1n4+iozONaLRIBM2zGDN6DMaMmYTWlk7cd+8DbglKy+3+9qLG8bEKXwumW7JkCSZOnNi3V1EUpX/oIgMZoMjZtGkTxo8f37dHURTlECOtb28DzPdv+JEYmukRvPLSRly06N1IxWP4r//6T6xYsRl33f0iguGxSEnblEo3IRJMoExER7IrgvZkD4JRLpcch5NKICoCJxrswamzazB8aAoF0Rb09OzEo4v/htPmT8BnPvtejBhdhZE1E+W8Ie5qa4lEELv31OL73/0hqqpGYHT1OHeZ5n379omwCKK1tdlNXXlpGQZVDsG27btQUFiAsrJStHY0iQAK4UuS1mAwLtmJi88exLtbUbtrJ557ZjW+/KWfobWxHOefexVOmHYG0k4JVq3di/seekYM2jLZFrETCEpaEojE4ujqakUkUoAwX9BJ9Yj46kAiVY9BFSF8/nMfx5NLH8F99/0R02eOwT/u+z8MGyYiDQlx4t/pXT1Niko5hvjWt76FG264wf2bUxGvueYabN++HZ///OfdfYqiKPmiIzgZ4NOjj370o31biqIohwFjgLvv3xgHPPXU82hubMOoEaMxqHwoVq3aIIZ+EZIiQlKJiJxWJsZ7OZrbY+hJVSAcqUY8XoyECKJgoATdXT2YffIsDK0ajDoRKSXFxXjm6SUiDqL44AevxBCO3AwfLOImimSaH90sQDRWhr/88W5s37wFM088EXtFmBSGQ2hvbkDYSSPZ04nKQeXYvXs3dsqxoqJiER1RNDQ2oWbMeDzw8GKseGW1CJwSdzTHQVQEUDGGjhiKadMn4b3vvVyy14KCWA+2bFqBZ5Y+iO1bV6GoQDoiEXfplDh3IbmIiK0wQuESyWMM8URSwkqhZvRoDK2sQktLCzZs3Ijp02YgFisSA3iXlM8mtyi5rLZy7EIhY0ZwbrzxRlx11VXYunUrNm/e3OdDURQlP1TgZGDBggXu6I1OVVMU5fBCUZMQUZASF0QyBbyycgPiKQcTx0xAa7uDHXuaEQh3Y+zoGMoKOxFKdyDenRABEJUzg4jLSQHE4KQLEBZxEQ1G0d3ZirLyYpQNqkBnvBut8Va87co3YdS4oSivHIxoaZXEWyoipULCLhaRFcEJM+aiobkTS597AelgEMUVFXBCERFS3SgfPBy765rQ2ZPE7j116JEwOzpaRXAl8MRjT+CkE0/GmJETRKsVSpilCIhLhwpQUFyCQUNLcO75J2Pm9EF44ok70NK+Gbv3voLKSgfTThgpBm2LCJq4lEAcjsQbCgGxYAIlkU6cdtJwzJxSjPdefQ5mzhgn8SaxbuN2jBo2DINLB6G1KYFXXtnIMTARaixPvs+TRJ9aUo5R2AdT6Nx666248MIL+/YqiqLkhwqcLFx//fXuUyRFUZTDBxd57v2Pyxs3NvVgzeoNCDphTJwwAXv31qGjsw3RSALvfvciXPPei/C2S8/EwgWzMWIkRYoY9EEH4XAYEREjyWQc4yeMQWVVGV546VnUNezEfQ/8FYOqCnHO+QsQK4y678hAhAjckZawm4q0GJMXXfRmfPVr/4t9Tc3o7I6LmEiguyeBbTt2orCoVMSXg4pBg9HTE3en8QZFjKxatQpTJ0/FV7/yNQwfJmLFFRbStYj4ckR0IRDD4MoqDBlShsvecQFaWneiYd82XHnVW8VwXYhhwwajoDCIQFBESTAluemSPLQiHOzEWQtn4F1XnIM3XTwP0WgnasZUobg4IuKoVsojjKFVQ9GTFMGzdrOIPMlSgKXJ+EXp6GDOMQVnTJh3YG1HOF3c79ihdJwe5wfTlemYoihHLwNe4HBo268xy+XKy8sxQYwL07AOGTLkdX68LtMLkU899dQBvSzJ8xiuoigDHd7nlDd816UJO3buQDQYwogRw9Agxnwg1Y2K8jDq9r6Kpc/8HU1Nr2LwIGmXKovkLI7h8Ns33SISaNinMGrUEEybNgFTptYgWpBGU/MuLFhwEgZVFGOwiJtguFTiKxJXIGeyG6C8SksYCZx11lkYOXIkYrGopKVVxEyP2x6mUilUV1f3LvlcVoY9e/a4oop/n3LKKe45yWRC0pCWdotBRtzwuToav29TMagUs+dMw5hxVdi8Za27zPOf7/kztmzZjFQyJWFx0KVNzu/ApMnDMXR4gexrx9/+8Tvsa9yO7VvXIpDuQkiEUHP9XndKW2VVJUKBoISxFV1dHLkhve2xCpxji5/85CfuiA3hFHEzVY1/E7Odj2O/TfibbZ8ddqb3fDiCZB8z/X0md6Sm09mLJOVKh7GLck3Dp93SX9uF/u100I7x4hWzXgFpbB+vy5QWO7z+pNfE45dGG+Mvl9ClH14Hm0x5MU45fAx4gcNpZmy8br75ZvelRdOw0XGb++19xnGeN38JG8H6+vrX+TGOQujOO+/Exo0bXf9e7rjjDtx00019W5mfVNGZG4h+zjzzTPfvY5H+NLb5Nh4H0tjaaciUjkPd2Np59zZ2fnAaZK5wmSbbj3FMuyFbvbL9GUyYypFm/2tQV9eI9vY2lJaUoLS0GLt2bEcgmcTg0ijGjxmEBfNPxMQJlQgGOtFQVwtRHkim4u53YEKhgPz2YPee7fjTX36HNeteQO3uDSJyEphx0lhEo2mUlolYiYckWoqcQgSckKQgDSeddEdkurq68Oqrr6KtrQ27du2Sc6Lut2o6OzvR3d3tbvNv+qXA6ejocBciSJspYQERONK1BAIicNIFks5ChML8MGgUpeVBTJ85Qc5pRGPjbkyfMRVzTpmNE6dOkWyIQAt0oaCwB6fPn4GhQ0tQ37BN0lGH9etflHLYiOFDB6G0qABdEmdzSyuGDx+OSDCK2tpaabNFHIGLC7A8pd4P/O5N6Qe0BdjnL1u2rG9Pbtg28xwb9vncd9111/naAv0J/1Bh+gyTDtojTEumfpdT/rz58sK+q79TA9mfPPTQQ6+lg/YV7RhbQDDchQsXvuaH9hUXlvD2u0yf8WOc18YyQo2/mfxkguflY2MxbNpvLM9s8BqwTnhZunTpfqLaOObPiG7l8KA9wBsAnwDNmzevb6v3SRUbIGJXePtGt/0ca/Snsc238ehvY2saPrt82fh403EoG1sjhIyopnvggd4lbDPBDpQv0jJecw7Lj+F48RPjrCc23uNsQJl2259Jp1mtSDnS8EEKr3dvc8yPWXZ1daNExE0wlEZne7scSaOsuADLnn0ctTtfRSrZLgb/NnS1dSEUiLqrjwWDASQSXSiT82bMPAHnn3sWzjnnDFQMLkQ0FsDo0VWIFYgACIqgCRRI/YhJ1DH5FTGCFMIiWEhhYSGmTp2KF154wR294SgOBU1TU5O7wABHbihujLApLS3FSSed5O4LuIslUOiIwHDDjUocUUmvg4KSQsQKKeBHIZlux5gxI/D88mex7LmnUV09XARSt4izbhFrHe53gM46eyFOlXbzlNmzUD1yBGLREJqlbIpihZxPh7bWNne6XFr+bpW/Kcwoa3rL8vX3j3JsYz8I8ntwZNpNurvvvrtv7/6wbeaqbPny5JNPuiu55YL9BdtVGve333573943BpYL23m7r7nyyivddv+WW27p2/MvWD65RADLl9Pz+Q2ufGEZsB9jGRs48sV08CEvYRmde+65bvoMfNeKfdttt93Wtyd/aBPw3Fz9rB+mf8wGbQXGkSt81j1eez+4SIZdJoRlxXrl3a8cWlTgCHbDaD9pOBQwPN7g+VRk3uj0eyA3+tFCfxrbfBuPA2lsGZe38aKRT4FjOr9D3djyaRDFnFd0ZIP5YnyM12DKwyuyssGy9Csfv/I1L+8eqwJ6YNIrcCgNWlva3elXkWhY2iRHREa7O4JTNbgcC04/FXNmn4xRI0fh1fWb0dzc4YqUdJrtF6enUV6k5Vy+0+KgpbURoTCn3BagpCiGoqIC2R8WjRNFIM2JbRQhRgxwalnvezXsfDktl9PPJk+e7I7gUNBs27bNva84ijNq1ChX4MyZMwczZ850Q+D5hPWLAzqO+1JMwJ1GJvoLkUgQlZUlKJa0PLP0Kamzo1BUHMWGjWvFj6SGIq0njvvuexRPPPE8fv3bu7Bu3WYJI4z16zZiz+46VA0aCqRC6O7sQYGInVAoLMKO6et2cyLSnv9I3G5SlAECHwSZBzfEHpVmH8F+3BxfvHixu5/tPesk+yViP9zKBz6ctPuHTJj42M+b9pb9i7ErjDPp8B5jW09nto2AM6Ium11CEcb+1QvvYa/RzTLKZdSzXHmu3SflC6ep5qKmpqbvr8zs3LkTDz744GvlYZeJwdgal19++X7++tNv5oL1Jx8xSFGbycbz2gO8lqwv+dQr5eA47gUOn2Jfe+21bqPIpw80UjM1gpmmAGVrMPjkIp8nQLngTWTi85tuZKeHzgv38cayG1E2dl68ebQx52brJPrT2ObTeBxMY5uPqDxUjS3LhvnhOwq2P9OhZYJ1xy8NDMt0moRPgVhX7bDthpx59c4hZ9nZUyOVo5WAO4bDXzbIFAUpUQe8+8JhCgO+0N+F7q4WLHniMTzyyGI898wLiMfFTzDiihWKEvf8UMg19rdv24EtW7Zh167dWL9+A4oKSxEOcHnpEAIiCFwFIHXIxBuQOAOBEJw03+VJYMyYMTj99NPdekXx8u53v9t9N4f1+x3veAfOP/9817Eenn322a9NY+NCBXCniRnSEge/75NyFUcwGEVBQans40ppAVRVVkgc1ago52IHaYSDBQg6BWhs6MbzL2zElClzMW/emejqSmDhwrMxePBgST/LJ4WExBcOS54kDalk2n2PxxU4FFby+/pWUDmWsftZ9gl2P8T22V4QyBiVZmSc4oj1hP7sdjTbFCW23X5TjgwUPyYchuvFPPAyootpYTqYbvOQycwqYNtt2m/uz/XQz4bh+fV1vFftMstliBP2KXxQdiDGN9PvDZtpY9lcffXV7jbz7Rc2HyjSDjPQjyk342gr2DYLbQ2WHR/iGT/Mr98MjAPhcIlB+rfrqnL4OO4FDhsY07CwkvKG4XxuP8wTcp5j33g8JxP5PgEibFDZGNg3OmEDyhuCcbGRZJj2Ex0et6cv8W/uI+YJEDGNufHHm9duCNgAMmw7HD8RlI18G9t8Go+DaWz9RlFYZoyTT3zIoWxsjRihoDB+TIeWTeSw7vhNmfCWDfNjx8/jbMj9xC5hXikoD6TslDcW95qKsDAmeSpFC713BIbiwHE6MG7iYMw+ZSomTBzrTgmrrBqKS978ZgwfMQzxZI+cGXTPS4uhHxHBww9dFhaUYOKEqRgzeiI62/kB0AgSPRKPKACG7AT5Uj7HjCRmiioKhGDvCA7bDN57nKLG9xFDIpwqKipcocN6H4lEMGPGDHcUh/c2YT5CIpKYdjcvAQlbxFkwJOGGgnC4cnMyiHi3HE4FMaisDO0dTSgqCqG8vAihYEjETUzOLEE6XSTnlmJXbSPu+dt9qKgcjHgqjraOZkSiAXc6G5u13vd+GBfT7f4l6eh1yrEJH5DxvQUDDUJi9ys0brNh+nAzokCh4e236dhGZ4JiwBjmftjh+T3U435OTee9ROcnptiu834yDzB5jo3pe/pjQPvBvjSXIc4+gw/SvA/KDgaKD5ZTtnhNH5YrXva57Pfs/pTb9kNS1hFea/aNBwPTlEsMMh39tU94Dq/3wV5PJT+OG4Ezf/78/Z6ysKJx24gWQuPca1ja8AbMJma8MA7e3NkwjR8djWGKCu+Nbjdw3iFgNlxMk32OSSeP8eYzeWJjbvuz82KetNjTluiX5xoxxW2mJdtNnw/5NB6Ho7FlB8PyzRbvwTS2xH7yxmvG+LI9raGQYmPM8jews+O1yZZOHjNi1z7XQKGlozfHDr2GTa/AKSgoQCgcQjzeI64bp82fjtPnn4D6xm1AKInm1mbs2lOLzVs2obm9EQgnkUwlxMDnyEmv0HFE4BTEipEWIdHdxQUCEujpSom2EXEgQoYfznS/FRMwK49xpIiiqrctIhQvfK+F793E43E0Nja6x55++ml3qhpHNimGOHpDuOBAr33WJ3AongIJ+Y2L682fk46gtaUHPd0plFcMwvhx1ZLfACory0WASYr4gc9Alfgsk7zGsHXHLpSUlWNkzSgEIw4SIvZa2ncjGEuipKRY8tXtxhmRNBQUxCSO3ph6/1GORdiG2qMsXtivsd2z21X2s3Z7Z0SRgf0g20lve50J+mUbn68hyqnOXph2tuPsJ3h/ZxJTzC/j8hNJ+cC+wK8P4P1p9/FmFMs49vcsR/7NMiVm2zheB6aff5tZC+wjbT+mz7RheniM/Zvfw0YDjX1T1vnA/NgPBO38GcwDl4OF9oKdT5aDqZemHtn1lI7lx3Ll337lwjrrrZvK4eO4ETjG2DQVkTe79+bgNhuZbI0ab8RsN6xNridAhA2f7fpr0FME2CMJBu7jsXwxT7y8jSBhQ5kv+Ta2+TQe+TS25smXcaahtqFY4jGKt2zle7CNrd91yDUFjulhvWRYJg9sBJlPLn6QDSN2vSOOzC8bWR29OTZw31MRkz6dpvCAGP7lCEcj6BRxkRIRMmX6RDS21MmxBAoLI5g4oQYnnTQFs2efiMEVpQg4aXeqFpdx5vs1iVQITU2d2LBhJ9at3S4ipAztrQ52bK8TYRFHykkgFYhLeL0qwH11R9QF3+HhN2ykCro07KtHIhFHQ0ODGzZXUJs0kcJ7rIiLIjzy8EOIRWOo28slm9PizIgQp6QxdK7NFkRCwkw7STeszrYe7N7RKPEUoq0thV27mrBq9QYseeppd9oZAhR2Ir/SXDMaKCyI4syFp2HFi0tFbDVh+HB+g6cLkVAQxcWF6Gxvc0dxCiIF4pff9ZGYOarDrKnIOWax+0T2owa2bewH2Jbb/TT7ZPY9pg31a4vph/2beWCXDb6n6Z1JkQ2243Z6zINHpj/bgyrCtHK0nf2On1GcC/YTbO+9sNw4gkKYPrtM6WjrmFEok36vH9M38W/z8I7laPvx2kMm7+zDMvW3pk+mKPWbjsdjfmKUYZo+lfnmttfe8NoaB4I3j3QMk+XBv9m30nn9sDxZrvzbWy7MM9Or/fIbx3EjcIj3JrcbQfPU3O9mOxCMoZxNLB0KOALlZ5Bznz06lQtjLPMGtMuIrj83ZD6NbT6NR76NLX9tP94GlR0GxRSPZcrHoWhs+QTPL98UQH6drY23XprpGSa9Jn3ehtw7DcPA975yjRwqRwkUF/Ij9rpce37JH6gaUoXCohK0d3agrbsVW7ftketfLoInKEKlCz1drWhq2ImnnngIgUQSIScCR0QNV0VLp/kuTAwjR01FKDxUhE4YZ8y/TMTFIKx8eRu6Eym0djaKdugRoZNGMtE34hPokZi5jyNBQTz08P34+9//KvvSIkRaULtrB7Zv24L6vbvx0vPLsfnV9WhubMDkieNw260/wYvPL3OnsUHOB9rktDiSjogVpxBx6WZ6JG8dnZ3o6XCw8sVNGFQ5CeMnniHpPAVz512MyiFjkZC6n3R6EAxLoSTjiMr2wlPnYNqUsRg9shLhYApDq4agqy2JwlAZBpUVS3pqRdCkMLhsMEoKiiR+EWqukwKlzpEwlIGD6Rf82nK7X/Eal8SMerM/yCVyOE2Zsz4OFPYL7B8MjI/xejHihmnjLx/q2Q/p2O+w7c+WXpYF+0Xz0I8wDPZH/Fj5GwnTyunavAa2sLOFG9Pp1yfbfigSKEZteB73m3NMvu3z2EfyvP6I0zcK9utMr/LGcVwJnGzQsM33yX0+8AmQMegPJ3yfhA2pbYDzhue+TE9P/GBjRKPYbiwIGxXTuLLR9DO0bY6WxpZpZFoJG1ID99sdyKFqbE1Z2+ex3Pi0sT9D0gyX59jTGYxI9l4bppuiz/uEkB1lf8StcgSRKspX/ZOpHqmvaREYSQyuLMegweVIJBPYXVsvx8K4/4ElePzx51FUOlIE0CQUFY7AOWe/GWPHT5UKHpZAuOKaCJ10REROGHV1bZgyVcTBzDPx6uZG9KQq8NTTG1HfkELLvi5QQwREhKST3bxB3HPdr9cEQmhrb8Ndf/yju9DBzt17MGjQYOyq3YMhQ4fh1Y2bUD1mDEbVjMHgqio8vfRZvLxqDb705a+go6MTaY5G8T0c6VlCkp9IoBtRESYhuRebG7uxaes+rHq1AQWFE7Hs+Vo8tXQTtm7tQFeXiLMUBUqRnFqISLBI8hKWvBfgnj8/guKSaol3urvd0Z6QNA1BtKAQtbt3u1PrRo4eisLSAiSkHLlSGxcscAfGFMWCbakROZkw/Vumh5PsV9ifUYxkenjFfoH9A/3S8eEZ+xjuY/9o9rOfNquLml8zk6E/GNvFhMswvCLjcGP6fJaNSYdxZjaCsQWI148N7THaI/Zx5sX7AJL5pjN+WLa5ZmoYjH/C+sDtbLbNwcIZNW+ETaj8CxU4hwk+xcg0PY2GqjGYeVP5PZ3x80MRY9+QbCx407ORpF/6o2PDa4x6v3OM8c/z2BCaholPvRieCYeO294RglwcDY2taUi809xYFubJ3KFubJlHOz6WNztTWzj5wU7SDpfheDtX7rMbcrpMDTmva7anj6xLPN+uX7lGmZTDh+N+oJMiBQiL4zLKNWNHIi4CZ8uWnRg75gQx8IeivT2Mzs5i7KmLYsy4M7F+YxueXb4aqb4X+7lCmSt2UIDV67bjoUefx70PvIAnnt6MWNFUrN2cwCOPrEM6XoKulg6JK4lYsMcdGeGqZg5FUjCEe++7F7UiaC655FJceP5FuPTSt2H4yFEoLCrDjl173AUOBlcNwajRYxCMRFBQXIRxUn/+fM+fJZyQiJwIM4UwuhBxWhFNd6KjsRVtrUH88S9L0ZWoxq76Iryyth0vvNiARx59BRs2Nkr6ByMUrBDRVSR5iorYK0ZjcxAnTL8ALa3lWL68FmvW7kFPPIhhI6sRTwSwr7EFSSeO8RNHgQM4vSusJd2RKBU4AwMzatMfTDvq1+fY4fmFnWt6Gv0bZ/o6P9g/GH+mr7D3GWdGm7wzGwj7Dv6dSWzZeMPOBf37jXTZsH/Jlkcbv7wZZz8I9DtO502LtzwypZXps/3l6m8N9jnGZbNRGE8u4cQ0mmvthccy5UE5PATkovav5VByQgFBAzvfhkFRlOMTNr+pVAIp9CAQDmL9+i0oKRmGX/70T/j6TbfglJlz8fGPfxbf++7PsXljPc45+zK0tgK1e9uwc3cr0ihCKDwI3d1JBEWo9q7GlgZCKXClsWhUOm6uZAbx27MaU6ek8T9ffidOnTMCQ4cPASIVQLAU6WBYzhKRhChu/O8bkUyl8G/vfg+6exLYubMWu3bvQXtbGzasX4fhw4Zi2vQT3XdyuBAB39Hp6urE6tWr8c1vfl1EiogbEU8BiR9okXibsHXLPjz2xE7895fuQGvHeCSDUyTzRSKuY0gmJO0iTNgTpZIxObfU/Y5PMtUifjowuILvJzVgXM0gyV8LVq9Zife+910SdxV+8MNb0NK1Gz/8yf+icnhQymcOSkVwBUToBUUkSQCugFeUfLGnjSmKcuyiIziHgf6+oKgoyvEJFxZI88MtgQjCIlbu+MNf8M/7H8TJs09GYUExdu2qQ0dbEpMnzxShUoSlz6zBSyv3YMfONBKpKgRCVYgnOSUr0ucoVPhtmGKERbwkUS7HByORHoZwwXis29CKP9z1CHZt24fW+npRFB2SijaJvxMB8P2bAKadOBXPL38O99xzNzZv3oCmpn1IxLtRLMJh2rQTMGZsjYiaerS0NOHxxxdjyZLHJczf44ILzkU0whEkvvzCd3FE4KTiaNrbiq0bmvDLXz6Apo5BQGwsUgFJd7oM8VQRnFAJkk5E0hgQsRaVfAbRLaenJA/B8BA0tETQ3lGG1WvrsXb9bhSXVmDK1KlYu/ZVdHZ1Y9jw4Rg6ohI/+MGP8eqGbSJsYlKmkgouy6Yo/YCj2xQ4Km4U5dgn9BWh72/lEPG+970P3/jGNzBokHTmiqIoGeDKZfxGDAJRNLXG8Z1bfi7ioR3vvOLtuO/exdhX34wRw8dg0qQZWLZsNbriBSIGKsR4r0Q6UIZkKkQ5gbCczxfrHY7iiHHvBELiet/JCaBQ/uY7Ot0oLAxh/boXRYikMHH8aIRCaRSU8SOboijkvGQ8jupRw9EoAua+f96LF19cga1bNouoelHEQ+/CAjt3bcMLLywT8bMRzz73NNrbm/HZz34a5553lsQXRzjID3qmke7pQGtdA7ZtasIPfnwvnn6uCeHYdPQkhonwKpN0SvqcgOTDcc/rHWihQIpKnjjVTcRaOij5kjxwYCqYEP+dGD9uGBYsmId77/s7aut2Yu7pp2D2nFn4+S9+LeJrPOadOlvOc9yFG/hOkaLkCxcH+Ld/+7e+LUVRjmV0BOcwwKlp+gRIUZRccPpUKikiRwz4dat3YcO6Jrz0wnZ0d8P9SGBXvAsvvfgiaqqrMbpmhAiVpCuI0hQBTkw0CadhUSj0LjTgNulBLs+c6l22WQREQI6HROSk0+VIxIfKeePxm98/h9/98Sns2t2B2u07kepuh5NqRzgSx5ChpbjxS9fjq1++ARMnjMLw4fxezUiMrRmO6uohGDWyCtNOmICxY0bgnIXz8aUbP4e3vPlCFMbSImy6JM5OxDuaUVe7FxtebcBPfnofHnx0I2LRKejsKpP4i1wRFgiw+xFV4y4PLSJHNt3Fz2QX3+Wh2OFoFEelUrLNd4y4JPSpp87Fnt07sbN2pyuOzlx4Fp5fvhZtzVEsfWojWtvSUkY8l9JPOZTY7wvm80I23+80/v1WolQURTlcqMBRFEU5QnAaFc1w2vjPPLUejfUh1O7qEkN9FS699GJEYkFs37YZu3ZuxbnnzUc8xZfqe0S38KVYETTpiOs4PY2jHmJJuu/ABCFCyEkgIIIgIGoikBLRkypGKjVMhMKJ6EpMwq2/ekLEx1+wbfNeNOzeg67WBjlPlFVPC5xEKy66cCGu+9B7cfllb8Lb33YJ3vym8/Ced1+OD3/wvbj0TRfg0kvOwWf+41pctOhsxOPNItRaJK5OtDftQu3OLXjxpTW45Qe/xd1/f1HinIKunkpEI8PgpEWUSX4lmZQ3LATZYNrDrpjhyA1cgdPbPaUDKTmURjzZLuJrEObNPQnPLH0cbZ1NGDliBKZNm44lT64WXTcUq1bWYed2SYcEzPeIlEMHxQrfLeV7Y1zgZNmyZX1HMmO/eM5FTQ7nKlWKoig2KnAURVGOFBQkoSCamlJ4/NH1IkgGI9FdgPv/+RRmnTwTs2bNQH3THjz15GOYOXMyxo4dJkKmUwSOMd5FCDh0bMqDfU/LReCI47LTrtBJxeGkZBtFSDmV6IyPQFoERyo1Dr/9/SP4wg3/g8ceWIy67duwd+t6xLuakexpF8HShalTx7uC4qwFp+Hcs8/E2DEjMWrkEMycPhmnnXay/F2JtAgPLgbQ092C3Ts3YNuGVbj3b3/FF/7zf/D4ktUIFY6XVIxEJDJS0lGIYDqMUDrgpsldFEHEjZSCpJWrr8X6HEWOqBTJA4JJET1dkuceEXlnormpHq+sXIFEogsLFs5De1sXVjy/FrHwEBFWHVi+bGOfZqJIkj+UQwJnJpjlfrlSVb6rVRm4oqR3qXtFUZTDhb6DoyiKcphw0l3uaAK/4t87I4tGfRoOP0Ypxjcnk3E61oaNTfj+D+5GV3cZItFB2LVrNxaeOQ81NVV45MEnUd+wDyfOPFEExgS8sHyVhFGAaIjTzlLuezSBID/U2SnCRkSNxBfk+zi988D6pn2JUEgHEQrG5LiICTkWQhw1oytF0DThofvvxc4dO8V7CGkRHul0wk0bwyssiKKwsABFBREUxIKyL4GiwjDCoh/SqQRamprQULcPG9ZvxFNPPINf/+L3uOvOBzB85Al489s+gto9ITQ0AKl0sfgXcSMncmlntwzcd2QYZlTii8m2OD53Y2FJ3AEuNY12EWn1qBlVgvdcfTEeeODvWPnKChSVRnHDF76Iu+56AqvXNIPLTMcTCVQNKcTZZ5+EWCTVVw7xXqEk8J0eDuxwWYaA5INLdPemQcnFP//5TwwePBhnnHFG357+wfdb+LHDtrY2TJ8+vW+voijK4UGXiVYURTlMpFNtSCQdxGLFoKTh33z5nZOxRKWAK6ilQwX4/v+7H1/4/O9FrEwSYz4mhnod3vaWyfjG1/8d7776Wrzy8mrMmXMqPnLt9bj1R3/BSy83Ses9AghViFhIIxRtl8gkwmRIGnWO5PC7OBFwAeo4knJM4pVwQ6J+xMQXe78ZRbF6nDl/CM44bRBeWXkvnnvuEfrE5CmjccaZczHlhAkYOrwSJaXlInJKJbkBRAtEqCS73Y+QdnfH0dnZhW1bd2PN6o14/LEl2L1zDwaXVWH+/AvRk67E0JFzsXJVKxY/tlHSMxw9PVERODEkg3EJR4RWQAJ8bcRG0s1MiChLc7qa5CsaaELUqReRU4drr30LBg0O4js//D5aOhrwzisX4aKLL8d1192KngTfLZKcpWsxc1YEf/nLlzFqRKGcx7D4TRxHjkUlTpFRESk2ij5XfHLES9Kg5IQrjPHbXgdjMnCKGr9FpmaHoiiHG2nmlaMd82FOfUnzwGHZHYsfszQf5eTv0QbLU6ecZCcZj+H/fvln/OCHv8PatbuRSLDJDYm8CYvBLQIkFUFbh4PFi1eKGCoVo5tfq+S3bcpw/4PP45XVO/Dp//gYCgqiWL36RTzzzKO48so3oayEQqlNDMVm+e1A0OnGtBPG4KyFs3DSSWMQCXcimWyQ8Fql/nSJNd8l0qF3mhe/ixNP94h4GY6zz12EcLQMM06ej/e9/xN405uuQHMzP1h7Fz7xia/iU5/6Oj73uW/j5m/9Et/5zu24+Ru34uZv/gj/89Xv4otf+F984mNfEWH2v/jN7fdImkbgTZe8C3Pnno/KoZMkf6Vo2NeD8ROm48RpJ8s2EI1QTKThJJLuB0bDwaSktUfS2IloWPaFOKqVko4pgVCgXVwHelL1OPnkKZKvqfjr3/6K9vZ9qKysxJVXvdtd8rqtPSVlVoKglFk4MghbtjTj5Vd2S3wSSiog+eXS0zERVwE0NSXx+OPr8aMf3SmC7HlJCxdpULyYdsd2FDfEbJsPRHuxFxbwOvPRaf5dXl7+uuOZXKZ2hvv5weZjBfZDzI/9LhLL62jJA9N2NPY1xwosP/taHqt2x0BBBY4y4GEHYr7aryhvJOGoGNBOCf77xh/gkos+jHe987/xve89jGeXbcO+Jo68hPDqxjosXboDwdBQ8VsqrXKJO00tIYLnlu/ejZknzcLFF1+A7u4WPPjAPYhGu3HZ28+WzrRFxMA+OMl9mDRhFEaNKMOa1UsRDDaJGBgpBn49Ak4LCkJJEUH8Lo0IHcRFX4mQiISwp74RTz3zIrpTxXj+xR1IpqtQV+fg7HPehWuvvRHz578VI4aJMEkMw31/fxmPPrwBL71Qi9Wv7MHuXXyvZxDG1kzHFZd/AFe84/04+6y3Ycb0c1BcOhEbN7ahYvAkPLJ4Be76033YuaveHWFx5D+OWoVCIUlTGqkkFyeoE6HT6o5aOU6HK3yCku5IqFnSVIcxo0px9bsvwAOS9/UbX0ZCxNw1/34lNm5uwsOPrkJB8QgRjmERdIUiKEvc1dQee2wd4pLVrdt68NBD6/H97z2E97//27jwgk/g6qv/E9/4xs+kfDglTrtAP8wX/r1u0aJFuPnmm92/M32x3SwssGnTJnebv3YYFDlcpKClpWW//V5nn5/pC/C33nrrfl+XN4IokzuaHhKaB5cPPvhg354jhxFeRzP2Cn7GHc1iTO2OI4+27scAXHKaDX5/X+o83PBJxbGwKg47XHbKxyLG0ODv0QZfOs5keCh9iM1wwUXnYty46dhXDzz+6AZ89cu/wTvf+S3823u/j6985SH84feb0NZCUVMt17oc6VSJGOuyHRmKFSt24y9/eQaf+MTHUFMzFrW1W/Hb392K00+fijNOm4CCcJsIgmY48RZ0ttRjyuQRKCpIoLAgjmlTqzG4LIpUVxtioQCiXHTNSbjvA/H7M93JEJ5+bi3uf2glSsun48mntmLkqLkoKJqAzVviGDHyNJx19vvEsLhAZMk4FJXMxPQZl4hQ+yDedMkHcdKMC1FTPRdDq05Cc1MhGhsiaG6Oiagbi57kEDz4iIiRVDmaW9Oob2iVuB3Z3y0CjO//JEXEdKG4OCl1ezomTRqMkSOL3NGogIicolg3Aqm9KCvqwr+9+xLsrd+CxU/ej67uNpx95pk49bQFuPVn/0QiLUIwUYxAsAxOmh83HYZwbBzuufslXHX1L/CuK7+N977ne/jiF3+Pe+55RdqrdkljF+bMPRWnnTZbyqHvOinHHDTKKbhs2FZy33XXXef+bTvuu/322/t8HhnYhzMt7NNNv25GtY4kTJcRlEcbHKWjmOG1s68n3U033XRUzSJgmozgPpbtjoGCChzlgKCwueGGG/q2FEXxg+/ATJhQiTe/5ULZikiDWyEG/EjU1ZXhkYf34uv/8yh+8qPHpQMfhXhPCdJpcaki+S0Ww70MKWcIfnbbvWhpSeILX/isGO9RrHjxGfzj3jvwnn97MyZOKJdedR9279iEVE87BpUWYNKEGgkviTL5e86sGa644Ts/XFXN/QinpISjKYmECJLWMF5YsQdPPb0Lzy7fi81b41i2bBe641XYudNBU3MJVr7cjO6eEairL0ZR0VQEQ2NEsBSio7MS7R0VSIiY2bChA1u3J/HK6iasfKUBS5/bgi1bO1C7uwupZKEIkLCc54h44js8bSJG2lFSApQUp1FREcD8M07AuWefhLmnTERBtAepxG5JdzPe/a6LMHp0OX7xqx9jX1MdKocOw8c/9VHc/efFWL9un4TLd29EFDoSsLhkshjJeAV27Izi/nt34JWXU2hpHiF5Hi+ibwxC4RKEosDV/3YpSkpZKr2LDyj+eKebHY7RBns6HJ/S88EZ/85l+D/55JO45ppr+rYyQwOY/ZV5GHMsPJRTejHihcLB7yEfRQRX9juaRuaUo4cBLXDMsKtxxN7mca8f43hj+Z2faX4x/ZonDXT0R2w/mfCG6Tdnk/vt4Vh2AjzPjtM03PY+OtNI2EO8BrtzMWk2mP10Jn+EcZvOh788btLGNNjnZcuL7Y9hesvb2xHZx+z0GLzlyDD9sMvB68ebfu9xxknn9WdfG8K08FxvPr15stNC570GJh7veV6YJhMGwzTnmfR7y9YuO3s/z7P9etNjY/JoMHkl9rXwXqfjiQDfIUmmcNVVl6C6erAY4CnX4A9iRO+CAqGJSKaHi8FfBUeEDb9pk06HReQUivEtYigwFA37QvjSl3+Bk2fPw4c/8mHZDzE0/45HHv0bPvaxd2PaiaPQ09WAl1Y8ixeWP4e21lYse24Ztu/YgWHDhmLBGfOQTrQjkk5KHF2IiuAJcqk1kT3JZAGSqXJs35WQxI7AM8u2YZUIh1fWNKCptRAbN3eiriEkQmIkurorsPjxDdhRG8TLq1vQ0lGOP/9jORY/uQEpEWgPPboG/7h/hYibjWhqCkk+SiX9pYjGCkRYBBCJitxLNaC4KIkTJo3CzBPH4ewzT8G+ui1wkq148vF/YPXqJSKYahEON+Hqdy7CybPG4//9+NvY19KAcGEJvvBf/4mXXt6Gu+9+HCF32ekyV9ggHRUjKCTlyHdqyqV8h0ucYyT+MZLfCQgGhkt6ouhOtOGUORNx3nlz0ZPokONcVkHxg20BBY13ipk9JexQwHd7OGWNYbPt4oMz/p1rRIHT0/KZ1fDQQw/1/dVrEOfzEW62WwfbR9F52z67jcwE46GfXG2+N31Mm40Jxzi7rSamjzAu03eN7Laczg7HWx50TJeNN510Nt50mj7H5MeeJWD80o/pq1gHtm/f3ufj9eER/pprwV/jjB9jp5htOm95ecOl814jhuO9Dvlgp4XOpNXgtRMyxWv7YxjesvdemwGPNCQDmiVLltAe6NvqhdvSoPZtOc7NN9/sOnLddde5zkB/3vPFqHek8X3tbzssnrto0aK+Lcf1Rz+ZYLze4zzfpMHET8e8EPo3+0za+Mu4eJ4dHvfxmEmjX34I47TTbaeBMEw7XFOuphyINy5ip8fOi50GlgG37fi88edKjzf9Jn2mzEwcdCbNJj1m25xj6gLhtsmPOZ/OjtvOoykD4+w0edPI8+xtwnBM/H5p9oPh0I+B59np8+aT8Li3PE1ZEZ5jH7ex82jSatJgnAnLex2ON+LpOqcn0e6k5e+v3/RXJxp+l1NS/GUnGvupg+Dvxf3ZCYTudgLBvzqBwH3iHhT3gGzfK/vlWPgPTij8Hacgdr1z6aU/dV55ucu5+l03OQWhuc6wikuca6/5P+eXt77qnDHza0558L3OoIJrnVOmf9d52yV/cd7x5vudty36p3P+GX9yaqq+6cw54XfO2OE/cmKBrzoFkW86hQXfcQpj33Fi8ndB+BtOYeirTnHkRvn7805J4Red4oIbnIryrzjRyBedwug3nVj4Zqc49iWnZtRNzvAhX3aqBv2nU1r4afH/Kacoer1TFLnBiQU/J+F/3ikMf8UNMxr+tlNYeItTVHyzU1b+ZWfI0K84b33zfc5Z8253zl/wa+eaqx50rrzsbmfROb9whg/+hFMau9YZVPhx58NX3+v84rsbnDPnfN4ZVHK+U1R4lnP99X9ybv3pq05llcQT/S8nHP11X9n9Tcrsfqvs/u6WXTAsLniXE47c5QQjtzrR4i86RRWXOXfes9S9NolEm5NKdbl/K/4czL1r2glv28W2xrSpxo+B+02bmOl8kq194vk8z3Z+Yfhh2krjDKYttuNkPNm2Cc8xebXbSBu7zScmrmxpNuk0YZuyMtfKhGHDbROP8W/OJ6bc7OvNfXQGuz33xklMugy2f4MdZrZ00o9dBqb8CMNjuZnjJl8mPLPfmx7+bZy37Gx/mfJhp4dxMQ3EHKezy9T2kwnmy/Zj0mPC4TG7Xpm00Z8dr9ln+7GvHcOwt48H9q9ZAxRWEHMDEO+F53FTMbwV2VQUU9l43D7XW4HNTWjO53E77kyY84yzKzTD4j7GbTD+bfz8eTH58cI8eW8GO19evOVEeI7fDUR/pvz80uiXJoZjl0G29PilxYtfQ+M9zy8OnmfnyZsu4hc2t73XPVP5cB/TYZx9Xj5586bJW8Z+afGGy+N2GAwzW5zEG67fdfSm5XgjnmpwEsl2J5VMOzu3tztz51wvRveHnILiHzqIiMCJ/MMVM0ERN0Ex0v/lZDvIY393wuHfO7GCb8v2R5yPf/xPzioRORdf8HknFJzrVJVf7Hzomp87v/npJuf80//XKYlc5ZQVftgZUnG9M6j0MyJy/u5c/uZ7nEvP+4PzuetWOVdccr8zf87vnPKyr4nA+ZoTiXzNKSn6litgSgtF8IiIKYzdJALif8R9WeL4slNYdIsTCX9X4v++pOUbIiL+2wmHvuLEYl+V7S+Kv6/K8a/J8ZucSPB/nfLot5zS6LdF4NziFES/KwLpG05xydfEfdY5//zbnY9f96Rz+qzvOEOKP+xc865/OFNqbnBqhn5CxNWVTlXZNc5nPvqw838/2OycdfIXJKzzRSTNdz7+yd85d/9lj1Mz7mtOKPIlif/7TjD0J7eMggGW3z/7yo2/LDsKxHuccEwEogihUMF3paw/4Lzt6m85jZ1xp7Ory0l2dTtOOtV3pRQ/zD1tO782zAvbEu95tvP2B35tTbZjTEOmNsXbHrKd8gsjEyZeO/x8+igD46Nf40xeiV84fu1zLhi39zowLSY93niJHTf9ec/35tvbR2TC+LOdIVMZGejXm07COO30mbTbaWG5GUz5+eXLG4dfmrx+7LKyMfuNs9NAvOEwXV4/Nn51zSZTOrx58Mbrd+38ymagc1y8g3Pttdfitttuc//mMJ5UBHfoncN8HLLjsJ4ZtuY8T6mQuPvuu91tzvOVSvray4l33HHHfvN+L7/8cnco3R76k0r02vmMl34yYYY9iVwP1/H8fPD6q62tdX9Hjhzp/h4MHBZmudnDm7lgefoN/7M87SHkAyFbenbu3On+5jP1IBtMP6+lHQenS/Bl+sOBGZY+99xzX7v2LKv+cuONN7rTNQym7pk5y8wTP7JnY+qIqTOsoyYMM/x9sOWpAMF0EZx0SK5tGkOHFuNjH38XSsvSSCT2AaFuuenj4tgMS30TP8HXXFC2w3IsirRT4r6PE45W4Re//Atu/829+NYtX8HCMxegua0Jf/zTnfj7fXfi/R96Fy699AKk0j3o6ulCItWNxuY9WLv+BXR170FXxx60tdYilWjEBefNw4yZYzF8RBnaOxsk+jh6EgmknIiktwBBfhA0VChpjEn8IfmNSvycalaOWKwKgWCx5KlE/AySnr4YkSin08l2UNKaBrrjSeZIwnIwZcpkvPWtF+Jtb2V71YGw5HvihFG48MJzsHHDOuzdswd762oxcvgQfPYzH8Wo6kH4/g+/hdWrX3bT9P4PfABnLlyIL954K3bvbpPzKyQdlVI2/IZOWMqJZef0lR9/JRqWqfyNQIdstIprwqjRZfjkJ96J0sIIQvIfv4njpI6LLvCAMS/F247toZmGk2lqzdVXX+3+ErY/9vl2G8c2htuZpkf5wfaJach34ZULL+T7b6/HTjfdwU7fMVO5+OK7yevhgvmnDWOnn+03y8a039XV1e6vwWwbP7na91z9KsNgvObbSHS0rWyYzrFjx/Zt7U+mdBrYLxrYx9EOM2lh/aM94OX6669/zbYjpp7OmzfP/T1QzFQvpsPklek5WHLZbLSb/GwCloPJo5KZ46J1t0UIBQsbbSNCKFhYaW2MIGIF4s3J881Nw/m89rxfVjSGxXB4M7FhpwDi+YyPjWu2hoQGNBuFQ7EalddoPVg419rczKL+3Rs8WyeQ6abzM7APhEzpsRvug4HpZ7gmDuMOh8BhWtkhLVmy5JDNaWd50LFOscwNfgLT27CavLMO877I5+VdJTdOUESBiAS5MgiLXrnsbafg7W+djXRiO4LJFjHGE+7L94GgGEOiI8RWd+FPUDZC4oJixAdRLkJiBBIO30l5BL+4/XF8+ztfw1veerGIiRYsfuxR/PznP8M5C+fh09e9F0PLwwglu7DyhedRu6MOVUNGYe2G1ahvrMWQoaUoKw2gONaN+fMm4ZILT0FJcSsKC1pRUNyBdKBFXDuSTrvE2YlEslV+O5BKybFkpyvAIsEIkGIaKRWCSMZ7EON7NuEkBg+K4JRZk8WdKMd7UBCU/WkRFZKPVDyNfQ2NqKvfi2eeW4KXVy1DMNyBsxbOxn98+lq0iiD78Y++h1e3rZKY2/GJz3wM513wZmmjf4MN67uRTo8UATUY8VRUyo5LcIsgg4jAQEgcP+rJwuO3dETsuO8ZJSSR9XINtuND71+ABXPHyGFHdJukukjOCUomlKx43wewYb9lt5WmH6P4YNtGsvV/hA8PuZxuvu032yc/4zYTTJNJA/Ni3q2w002Xr2Dyg/2QeV/Jtg8OF3woSwPbmwf7HSMjUAy2YKHLVd65+tVbbrnFtX0YbyaYzq1bt/Zt7U+mdHph/CxX+0Exbbb58+e7f7PsvfYF+zz2haxXrIcmrgOFopX946G2BXLZbMyX3ZcbWCYHm6fjgeNC4LAisHKykpqnCUaEULB4GzYanKxUbAx5U/F83sgUK34Nq3mCTvHEsNjA8XzGYT/JInzKw5ve4DU+eZwN5YFg8mk+yGawG3XTaNlCxS9ONg7m6QcxHVe2xoh5ZTjesJnHg230s6WHZW6ujw3jts/JBZ/+8Dra6effDOdQYxonuzxZL/waMy/0Z6eJ15tlbHd0duPHOutd8Y7nsK7Y/nj9aGzwvvBeL8Zn11slPxypt1J5XWM7EEihtCSAz/7HFThx6lAE0vsQDFBQdImR3iPGe0KEjqgct1nuVToBOcKFvhynEOkURy2qkUyPwI9/ei9+dtuD+K8bP4P3f+Aa9CSSeP755fjOd76JgkgSN97wUSw49SQ5N4FkwsGji5eIOKKICqKgMIad2zdj8sTRcrgFc2ZPxqhhpZg3ZyLefPHpGDE0gqJYp+wrxLDKGKoqIigI96A4mkBYhA4/vllcEMDg8kIMLiuR1MZRM3Iw5p82HXPnTMUZZ8zClMnV6OluRmVFqRgF69DW1Izmxhbs3FGLf/zjfjzz7DLs3lOLoUPL8Z73vRNXvOtSETyP4Ve//Dn27t2DgpIi/Nf/3ICZp8zH52/4GTa82oVwsEbKocod0XIgorHPrqKccQJB16W5U8o5GIhL3jsRiiSQStRi4ZnjcO2/XyTplyKQ68EV5VLBpAqcHLDvYJvI9pewH2EblU97znY5m/FroD8+5GMbZvqpbLB9MsZtf2Abxrwc6kUSiJ+hasqsv7AMeG42AUKbg2267Yd9nRGjbNvtB7f0R2PfjDqYvtruH825hlz9Km0p2+DnPu93X4xtZMfDPtX0JUyPV9wy/xRPixcvdre9xr/pJ1nmPI92lqmPJq12X8h8HCzsJ71p9Pap+cAweG1NPc9ls5l8Gf+Ef/Pa0V5RciAV4LjAb04it+UG69vaH6l0OeeA2kjjvF9YPJ/7vDBMe79Jl3Hctvd95Stf2e844+D59j4v9GMfZ9ptmDb7OOOz9xFvGHT0Y2Onw5SVNz92Xk0ZGsc4/NJib5vz80mPX1iEabP38xpmKiMes/fTmXDsfSZd3rAff/zx/bZNufhdM295MG57nzSg+x03aWRYJlyDN+90dvl48+stOwPDpl8vjM/k2VtGPOZX9n7X+3gjLi7FFQb4TyrlOEnZ1+M4v779Raew+DonUvgNJ1T4aydS8Eex0/lS/ANOEI+4LoQHxd3vhPGAEwrc74RCf3FCkdudcOw7TrTgBicSfZ/z9rf9P+eJxY3OD77zgDNy2MVOaWyBU1V0gXP5Rd90fv69Nc7XvvCsM2faV5zygmucWOhKZ+igjzjDB33UGVr+EefKt//RueT8XzoXn/tL531X/t259j0POV+5/hXnjFnfdS4+65fOf356hfP+Kx92zjn9N8558//gXPnWJ5yRVV9xzj7jLufi8x50rnv/FmfuzL84g4q/5Vx12fNOzfBbnEljv+tcc9VS53Mf3+ScOuMOp7L4f53S8H87NUO/Lsf/yxlU+kGnMHylM2LQh53L3vQL5yffW+t8+T8fcmZPv84pL1rkFATPdWZO/bDzy1+scW6+5Xln2PBPSz6vd6KxbzvhyK+cYEjKKfR3B0G+s8T3b+6Tvx8Q92DvrxwLhu9ywuHfyHk/cgKR650xEz/nPPbEdvd1m7SUP/9IOwm5Nj1yOUT2KVmx72HTBhwMDMO0ZV4Yhzlm2hn+Gvh3tjTYafVzmdo9g1+b5de22dsmPd421rvPG85dd92137Zp0805dr798KbDe443Pm/7682rvW379cu/gXk3+03Zmn1m21xH47x9l1+ZE55v8mPHY+fblL3pb4hd5saZOO19PNdbhowzU37tfSaNZtt7DrcZp72PeTFl4b0W3jSbvBjsY3SmXPzS7xeWXX50xwsB/iMZVhTlGIVPtDg66X2qyidF0rjl9bRVOTwkEHencAXTHJkB0ikgmXIQTwdx09f/hm9966/S40xGMFQpB/iRzzIEnELXLwJJcXHZlhD4HkywB06wXa5rI0KhJkSjLehs246ZM8bg8zdcjVHVZfjf//kenn78afecUdXjcc55F2POnDPw6oatePyxZ7Fm9Tr2bohGYiirqEBRQRHa2zvx1re8FQ17a1FRHEN3dxcKCgtQWFKKnp4k1q3bhGQygJNOmo3axp1IpoFVKzdj7pyz8OQTy9DY1IBFF52F7TtfRWdHM2ZOm4W2tm6sW7sBexsaEApI1pLtCEXiKC4N4tRT5+GcBWe5eXv40b9j2fKn0dHZimA4gisuvwyXvf1teODR5fj1b+5DKslvAw2WEhziLmfN7904TtQdGQsEuqSQ+H5TTH45NiMJC3QgFJSwxDnp3RJfPb75zQ/gve85FdFQGmH3MkiXJ2ly5L/ep+y910bJDtsZjgYfjhGQfDFPso9kGpQ3Dt6fuUxUjopwxIPT8zi6xHd3vPXjaKi7yhuPChxFOYbhcD+Ht71CxkwXWLJkySEZolcOjCTaxDiPiMARI9wJiIGeRkpcUgzyhsZOfPozv8A9f9kEpKrEzB6JdHKI2OnFYnpT0MSlhe5BQMRQyBU4KaQDnQjQgA80iRiqQ0GkQwz5ZpRXduLfP/BWvO3NF+PB+x7E//3q19hVuweRaDEmTDwR5593KcaOmYk9tXV49rll2CRGQV1dvcQXRCxWiIryQSiOBVEUBapHV4voaROBNFriDGHzph0IR4oQicjBggSmz5iFP9xxP+adch5WrlyHPfW1qKosxuQpIyVtSWzZuAMN+xpFQ4SRSMZRWBjDkCHlmDFzEuadOgvJRBdeWPYcnl3+JHbXbZUoUjhx2gn44Af/HVVDqnDrj+/C0mc3ikipEgOnRMptMOKJUsnzYCm/QjF4wixKOdYjx+RvVxBK+QQS4qdFBM4++bsB4fBufPTjZ+G///udIuS4AEHSNZgCRtDwvSdX4OhCA/nC8rvuuutemyL8RkNDldPK7am1ysCF4mXChAkZ6xzrA+FUOeP35ptv3k/IZOojlYGPChxFOcbxm/tMNm3apIbAESaFZjGo+b5IgTha5RxloMDh2zUhbNjcig986Kd49smNiIbHiFipQSpZJn4KxZiPIxAR//EEwgEa9RRH3XJeuxxvEbO8SUJpEUO+A6HIbhETTTht7gxc/x/XYagIil/98jf4y9/vQ3NLO8KhMtSMnIkz5p+JGTNOEiEQFqNgK15euQbbt+9EW2u7CI8epFIUUL3v6RQWFWHQoCoRSYXo6kqhs6sLre17RShNxLYtzYhFBqO7y5FzEkil2ySvrZKuuMiaiJxfKOdWYNKk8RLfiaiuGYrdu7dj5SsvYtmyp9DUuAuBcBJVw8rx3ve+E4suvgT33btY0vwn9HQHJZ/DkEhUiADjQgIVEu4gETLFkrYC+ZujNvLjlnBUfrmaGt/D6ZB8ibAJ7ZUi24Grrp6Hb3/n3zC0KijlGpfyogDih0A52sOzuUoC/9QRnP7Ap+T2ew9vFDRUOVLNJ/XK8UW+fZwRM170Qd/xiQocRVGUw4QjAgQicNwljQmnnSHu/pniqEwghhdWtuCa93wVWzYlXIGTTg+V/aVi3DvgzLawk0YglRJzXLb5An2wRwz9TqRTDXBSTRJ2G6KxBhE4zSJ2elAYi+P2X/0Il166AI89vhy3//pOPPrI02isSyAUjKG8rAJTJk/D7NmniXEwCUgH0dDQhO27dmHnrp2o3bNLRFEz2jpaRUiIOBCXloSk0yLJJP6QCJhUIiZ/F0taoohEgigpi6KyqhjVo4egpnokRowcjoryYjS1NGD16pV46eXnsXfvbrS3t4oY6pR4h2DRm87HB699D0bXjMJ1130Rd915n4RV5YqxeHKoCCcRNk4hQqEKkSOlsh2VEhBxIqIkKL9Bd2paVPZLmoLiwm1yfgM6OtfhwgtOwC9+8SFJR6E7QsTV3+Rk8Vckvyx3XgG5Fq7AoeBRFEVRBhIqcBRFUQ4T7miLGOH8nk3vdsKdWhVwLeswEqkI+D7Js8/twrUf/DY2b04iHBwjYqYSqaQY9KEoAokkAmkROOkAnGBADHlHhEU30k6TCBxOBROx49SJIc9lnBsw66Qa/OLn/4uTT55Amx6dnWk8/8Ir+OvdD+KRRx/H7p316OriaEYMw4ZUY9iwURg/bjJGVo9B2eDBKCyKIplKoqu7S/z1oL1TfjviEn8aoZQjnUYYsWgpioqKURArQGlpsQisEHriHYjHO1FXtxMbN63Hjp2bsK9xD1paG+TcBIqKizBxwngRXufisssuxuQTaxCKAN09Dn78//6Er3/9dslzheSpFD18H0lEDQVJ0CkVkVUgaQqJAJQcUeAEReAkY5IHKcmQCMZgh+yrR2fXelx00VT84PsfxbixxQhH6F+EDEfDAhJZWhxH0gjFJqepiWBTFEVRBhYqcBRFUQ4TXP4ZDo1qscrFrnYQF6M8IUe4T8SNtL7JdBLxRBjLl9fiYx//Adav60IgNFrESjkCkTKgJyUNNSdUceQhIsfYZHeLwd6GsLhgqBspETrx+F5UlKfwy1/8F976llkiKsQbvwnD79BwkEJO27BhL5579kXce++DWP7cy2jcx6lpEkU8hcJYCYoKi1FWUYqyQWWorKxESVmpiJciFBaUIBqNSaoLkexJI57sEeHUIoKiDS0tTWhubkJTUwva2lrR1iHCS4RcrMBBtAAYOrQC5553Fi684ELMnTsTw4fzvRqOYLnLAohwAWp3teFDH/whnnlqO0KBKhF2ZZLfAimyQikHfmyUU9HCSIvI4hQ/fuIm4i7cIOUZ7kASe+TYFlx08Yn41jevw+RJIo5CadEyFIPyy/EvEWYUZ73LS8s/7jtOFDu94lNRFEUZOKjAURRFOUzQsOaUKNPIBrgqmmvWi8BJcdpXEkmnQ4zvYoREhSx/vg6f/OSv8PzzXCBgPBI9hQiKkR+gMR+MSThhcfyvS4z8VoSCXDWsE2n5OxrplrgaMXvWKFx44SkiKuZg+vSxKCoKMBVwRBzEohw5EmGRBOr3duLpp5eL0FmHNWvWYv36DdizuwGd3e2SRp6Tln8lzUERBeKYj2CqpG9fAol0h+SHU+748n5I9ov8EYE0ZtwoTJ4yFlOmjsGZZ83DKXNmoqIi6oo56okUB1Tk784uYMVLa/DY4pV48IEVWLOqSdI1XA6WS1RFUi6cGiciMM1FBPgOEwUeV1VKyr4eFIhSiUbi6ORHUyO7RdxMwY9+9DERUFEpkzZmWMRNiQgpjtRIjkTcUGhS4AT4T1DEp5sbTnVTFEVRBhIqcBRFUd4w2NzSiWHdt+hAyhFL3xHxkhYhIUdeWd2AL33p93jkkc1IJYaLTT4UyVSx+C2So1ywICiip01+60QA7MGIoTHUTBwiImUNEnERT3xnBwkMLg9h2gkj8OaL5+Okkyegenwlhg0rRmHf60AmKbT/29s6UN/chL2NbdiyeQfWrd2E2to9aG1rc9/FSSZSiHMkKeWIeIggEi4SMVPoiqdBlUUYUzMakyZOwdgxY1A1uBSDBpWiWOLnoFOK6koy1i36q66uFavX7sWy51/Fk48/j7Wr96CtWQRXWgResFiER1i0SISvBaGzh8tixxCJDJYwOOpTxqyJHwqcDhSEuFRDvQjDbXjfv/ODoJdhxHARK0ERkcEuOSblFCgRQcTloJnhXnH3Lyg0iXe/oiiKcqyjAkdRFOWIwXESMdQ5OoGwO62MgmD37m58+lN/wz/v2yRSpVJs8ArxwylbMYSchBjt+zByiINZ0wdj586X8aa3XIQ77roHe+q63FGLtLt8creIChEAaBdhU4pxE0ZgygmjMHPGaHEzMHxYJaoGRVFUEEFMRE+w7/M7JCmih4Mc1CZxEU3xeFJEjrhUUkRDCAWRIhQUBMCVo/kezWvIOek40NUJEUYJNLd2Yueueix9dhVefXUn1q/bji3b6rBvXwKRUBlCIkCCKBIBwkA4nSyBeKId4XACM6ZPR824qXj8yVfQ2EiFMhSpJEewEgiH4iJymlBS2oJPfepcfPJT81FRLiFId8apaYFAn6rS92sURVGOS1TgKIqiHDHSIkZS7mgDp3nxvRR+SHPb9jg+cM1dePqpBgTDQ5BOc2pVVMRKEOFAt/ivQ0VxC06bNwKjRhRgdPVYPLv8FTy0eCnSIhycSAmSTgCBMM8RYZJOAN1tCIVFGAS6EI0GRPQMw5gx1RgzdhhG1wzByOEFqCyLYfCgIeKqUF5ehOISET+FQYSCIoBEg3GmHN+ZiXdxRMZBZ3cXWlpb0NjEd3Ba0bCvC7W7RNTs3INt22qxedNONMm+RIKrsUVEoAQk7hIRISUi5DhXTURSJIhkvFO2uxAJpzFl6nicc/oMlBXHMGT4eHzrll+jvjEi5VMtZVUk4kvywOllwUaMGJ3G3X/+JE46KYY0V1NzkpJWjoT1jsoE+ZVRRVEU5bhDBY6iKMoRhC/OcwqV+36KGOa79nbhP/7jNvz1z9vFWJ8moqCYM9mksebUNPnDacGgsk5cd+2FiIZqUb9nLeadfCpa2hJ48rkV2LB9D9Zt2Yv2eATBSAVS8QDCkQIUiPGfTLaJ8OlEOCpiIx0QscL3eRIIxVJyvBPFwQAKYxWu/2iBgwJ+IDMk4ijgUIohHUzIv1FJc6kIirAIlwS64x3o6u6WsBleGJ3xFByeIiImEiiWTIXBVc84uhMQMRcMuTJNzmtDNJxwl7euGlyAKVNqMG58tRRIArMmj8bK5ctQPmQcokVj8bNfPISO7sESDpeMpmTrRDraKGFuxoJzqvDLX3wKo0eWSzklJOTej6qyY6MoUxRFUY4/Ql8R+v5WFEVR3mhEaHD1Y6qYeCKJL3/t2/jNb/6BUGSUCIJKMeiL4bgv2PNFfhFDwSRiBT04a+EsPPPMg2hs3IW2lmb8859/x3kXnIETp43H2DHDUFwYRk9HOxI9caC7S+Jpx7iaYagcVI6WpnZ3elgoWCYioHcZZiTKke4Zis72crS2FaK9NYK99T3Ys7cHzU1R1O0JY099EPV1MdTvLcC+hhia9omg6SxGMj4YPV2DRYwNQjBcJAJpsMiQcjhJ+TtQiJDkMZDuEWETx9lnzhOBFkFb83bUVBfh1DnjcfKsMbj4wtPQ1VqHv959B06fewra27tQ39iGs867CD2JYmzd1irlI2UQCvd+DyjKaWgt2L5rBdrbduDss+aKOIv1jt5w9IpFqq/XKIqiHJeowFEURTlCcPzcXfnYnUkVQHtnB+64826sX78TsdgQJJJRBN2PU/aOSHDJ50Aoja6eVuzctQVnnr0AZ51zLv5x3z/Q2t6MouIIgoEezDn5BETSXbjonAUYO3QQygqSSPbsQdXgIC4873Rs3LgGPZ0JcCW3EMIiQiTcQJEIh1IEwnQxBEUspINB+btQZJWkIVwuwqVK0lApcVQiEh4siY8hFBJx5DCMEtmOSC5EeKQcBOUnmEpyrEb2NyEW5IhNC0YMDWHSuHJMP3E4JoyrwKjhxSgIJ7Fx7WpMmzwFO7dsQ2dPGKGCodhV34VlL23FvpYQdsvfaUQkjXxBKIFwOOW+h+OgARVlabztbRegpLBIyrRvWpo7Ktb3t6IoinJcoQJHURTlCMKvtNAQD4p4iUVjOPfss1BSWoHVqzagrZVLQVMIJRGM0KDvdt9ZcQJpNLc0o7sngabmLixdtgyz58zDaaedjrCEs3HdGmxavxYnThiH5c887i4mcM45szB54hB3hbEpE8eibs9eRCTe7s4WOIn23tXcIgkk020Sl4gJp0XS1oKKwVEsOPNUETdhd0SF79OEQmEkkz2SlgRSqQZEQp1Ip5pF2LSJBJF98TYURR2UFqYwekQRTpoxEifPrMY7Lz9P8tCIU06eiLVrnsfWTesx/YQTUVJUjjWrN+G0U89CWdkw3HnPk3hhzW5s292MbbVt2LW7yV2AISxhItAq5dWMePdulBR344MfuAQ3ffVjIpwovDhkQz9UjRQ4OoSjKIpyPKICR1EU5UjhfpTFEaNfrHH5k1/mLywswBlnnoQzzjgZnZ2t2L7tVXR1N8hxEQ8xEUSpuIieENLJJOrrmrBx43bEkwG0tfdgxvS5ePjhJ/DkE0/jrW+9HC+vXY0de/bg5NknYXR1JXbs2ISOzmYRGCdLnCmMG1eNyZOrMX78YFQODaGgqBvFRUkRKk2Sjk5J0j6MranESdMm4cUXlqG1M4VAlN/lSbripqSkR0RGB8pKOjGsKoDxYwdLWDU4edZkTJk8AhdfPA/Tpg5HU9MuEUU9OPGEyXhhxTIMH16FQYOG4J6/PIiGfUlUDZmEnkQhlj63GpGiSmzZnUBjRwhOKCaCrHcltFA4hXS6QfK/Q4RgAxacMQZf/9934wP/fh6GDy1yp6QF3X/FBXvL1Sw2oCiKohxf6CIDiqIoR4i0u0x0WgRDBKkEEHat9BSS6bQY9BG0tTp4+qkN+M2v/47HHl+J5mZ+MHMQCmJVYuiHpAFPy3YCATkxmWjDuDHD0d3WgLHjhuCyy9+MH/7ke2jvbsUXb/gsNr30DPbu3Ypzzz8LK19ehcceexZnn7MIiUQc4ydVg7O5du9uwKSJ07Bt6w4UFESxq3YHWls6ce7Zb8Ly5Rvwx4fXoitR4i6+PPukKRhXU4jiwhZMHFOGYLoHr7yyFjv2dqKouFDETwwjR1VKntL4x9/vw9VXvgejR47F7+78NYIiQC6/4r349a//ilWrdiAaK0cizRXk4kgFpEycCqSDIqRCXBI7LgUl++PtKIi0i3gajvddczEufcvpGDEigu54AmEJj9/nMVPS0hSN8h+XM1AURVGOP3QER1EU5QiRpkChIR7kAgIcwXGQTicQDPE3LSIjiClTq3DxRfMxd94MMfa70diwBy3NDXBEEfF9l1AwiBTfgQkXoam5A53djoiFCFZv3IFdda1o6kyjpGwIOpqaRUysw+lnLMSDDz2OE6fNwNQTJuHJpx7CySdPxqMP/1MUhoOygmJsfnWDKypmTp+KnVu3YnjlSGzZtEfCbIUTLJe40+hsacCY6kHuaE9PZz2mTBiJjs5WxEpL8NzypRLPXMycOROPP/EUXn55A7Zuq0c0WimiZyKWvbhOMjsUnV3F2FHbjXiqGEkUIBgpRIpf+QxxUYUkEj37pBwaUVjYhfmnjcMXPv9O/Nd/vhNnLZyAgmIKoS7Jf0K8BxBwl0zrFThOn9DpHdFRFEVRjjdU4CiKohwhOCWNH87kml+iU9xFBIJiqNPxPRe+l5OU/xBLYdzEIThv0Vycfe4MVA4JobV1N1rb9qK7p0nOS4nU6QZXY045AbS0dqK+oUPEQQECTil2bduLlpYWDK4ag+EjTsBTT61EXb2IkUgBxo0Zi5NnzcI//v5PzDxpLvbt60BJSQXOOPM0rF33Ivbs3YUTTpyF5S+sxfZdQCRYgmCyAwXhbrzr8ouwZeMqDBpUhFhhAM8896wIm/Px9JIVWLF8vSSmAmPHzJYwHRE4jVizditq69vR0OrI3ztQWyuCKc2PkqYlv5yC1o2AuLRTL8JlL0aODOL886fgi1+8Ap/69GVYcFY1ikvDkl8gHOL3eSJSRlEEApJxihkKm97/3TJVeaMoinJ8olPUFEVRjmLiaRE4InS4Ohj/46+TAnbtaMczT6/Hffc+ihUrd2F3XQu6OpLitVAM/yKx9QtEMPD7MxwN4TS4OIpiMZSWlrqLFyQSnQgF4pgzZxrOXDgPf/zjr/Huq67AC8tewao1a3D2wlNFSPQgVlCMU+deiF/+8u9Ysrxd9pXJeR3SeTThlNnjMHZsOQZVBDFlyijc+88H0dlTiWXLVyMs8fek0hhTPQXNIq7a29skPVxaLY5wVARZgl0Pp9lxxCUuxzoQCvdg8KBijB5bjIsuPh1nnX0aZs4aggJ+bFTUSiqdRsT9eKcIIncBAZUwiqIoyutRgaMoinIUk0o5SLsjPZzS5vSuJkbNIkKHeoHf0Nm8rRXPLXsVjz/2LJ575hURPy1IxGMihGLit1iEQxQpCh0RFGl+k8YVSo47asTV0ErLS9xFDKpHVKCrPYUdO3eJj7j4T6JiUBXOO/utWLlyO17d1CNipAhBdIuA6pBzG+W3DSNGlKJqaDG2bq1Fc3uxGyeXtg4EQuIniTBXfuO7QhRqqSYEnRaEQhJKdyeiBSEMGhTDtJnVmL9glrgZmDl9DMrLJXbJX9j9Rmha0sp3amQH31xKpSSOQgmfIkdRFEVR9kcFjqIoytGMtNCJBMUIBUNadENKDPyk++4NRzD47kmCpr/Y/qIlULe7G+vXbMZLL2zBihe24dW127C3vhGN7W1ybu+ITjRWLMJIxI4TluBlnwgFfsEmxNEhJyhiiu/1pBCJyPFUEIluTgsrgZMulzRwiYEEEEzIuV0SZpf47xb/3YhGw/IbkfTyvRiRIpLOVLoboSiXuO6R9HWhSP4eWlGAseNGY+rUiZhz6jicMuckVI8tQUkZ8yiiRsRMUtQN30lKJePg6zWSMrcMmPdwiFPTCt38K4qiKIoXFTiKoihHMelU2h2pCYWD7tLO7pwufueFTXffy/QJcLECIkLE/eimnCdip6sd2Fvbiu3bm7Fxx2Y8t3wFtm7djYZ9bajb14XunhB6esLoEb3ixEUwJIsQioiA4cd3AqIqRASFRLAEnMK+XwoYiiwghSScIFeBi4sI6mQAEn83IsF2EVCSkkjKfS9nUEUMpeVh1NRUYfr0yZg5bRImjhmOYUPLUTEohqhEm5LEM8x0oEfCiLvp55tJKQ5PSR5Dkpa0/M33lWSHlIP8HY66fyuKoiiKFxU4iqIoRzN84cZFjHmqAB+jPh4QhcLpWyJ+OHUtFAgjyJEZtu4idESFiBiRH9mOi7ctW1uwe+8+7N7TjZ0749izpxON+4JoagTqG+tRJ380t3SirSWOZLeIibSIJjfulOiNINIcTYqGUVRWjPKKIlQOLsPQylIMHRxCVUVStiMYOqwQ1dWF8luO0WMGu1POmB5OOeNnapgNpofT7pjutIglJ8DEJiWulAgq+uHQDfMhjks+u2nom5bW96MoiqIoXlTgKIqiKIqiKIoyYNBnYIqiKIqiKIqiDBhU4CiKoiiKoiiKMmBQgaMoiqIoiqIoyoBBBY6iKIqiKIqiKAMGFTiKoiiKoiiKogwYVOAoiqIoiqIoijJgUIGjKIqiKIqiKMqAQQWOoiiKoiiKoigDBhU4iqIoiqIoiqIMGFTgKIqiKIqiKIoyYFCBoyiKoiiKoijKgEEFjqIoiqIoiqIoAwYVOIqiKIqiKIqiDBhU4CiKoiiKoiiKMmBQgaMoiqIoiqIoyoBBBY6iKIqiKIqiKAMGFTiKoiiKoiiKogwYVOAoiqIoiqIoijJgUIGjKIqiKIqiKMqAQQWOoiiKoiiKoigDBhU4iqL0m4suugiBQAATJ07s2/MvuI/HysvL3V/jampq9tseyO6jH/1oX2nsz7e+9a2MxxRFURRFOTSowFEUpV/84Q9/wMaNG+E4jrv91FNPub82d955J1paWl7zs2TJEmzfvt3dNu7mm2/GokWL9tvn5zL589t/3XXXuc7eR8f4iXe/1xH69TtGlyktmzZtcs/lL7d/8pOfuNtebrvtNlx//fV9W7nxCieWdTaBZMSlcRRUNvZxrzhluOYYBSzPNdt+zj7fCN5Mzq4j9v7Nmzf37d0/fhO2nV7WOz/8/JjtTM74Y7rMPm952OEybUyr2fZz3rLOhB2u33m54qHLhPc6MCy7XO247LzTcTvXNfdih01nX2cD88t0KYqivKFIZ6woipI3IgCoBNy/J0yY4IiYcf82ePfRL8/xImLAPcbfbIiocERU9G39C3O+jYgb1/nhl1YvmdJqyJQWwv08ngmGyzTkA9PJtHjzwvP9ysz498bPNHnjpB9vuRkYnx2Gudbe+EzZ22XF8/zyxzC9YTBd3O/1z3x482z8+ZW7KVO/NJo4vHCfNywTvl/cps4wf/Z5jNMcy+fa9ucaEe7z+iXecvfidx1Mufqdy302PM593vI055s8G0y+/Mh2fRRFUQ4nOoKjKEq/WLBggTtC4/e0tj+MHz8eYvygtra2b8/rnyrT3XDDDX1H94fnE3sUIBsXXnihO4pkYz/xNk+3d+7c+do+71P9bJx77rnYunVr39brWbp0Ka699tq+rcwwP1dddZU7EuUdCeLIGcts2bJlfXv+5V8MW3z+85/v29vLAw884P7ySbvh8ssvd3+9IyIM59Zbb31dGDZmRIBlL8Yr7rjjjr4jmbn66qvdXzvNhCNZYvTmNfLBcnvwwQdfd60Zfz5lSswojClTb1g33nijm/9M9Xrx4sWvlacX3hNMR6ZzGVd/rpEfJv35lruXsWPHunXqmmuu6dvTP1huTD/z4S27TDCdt99+u/v33Xff7f4qiqK8EajAURSl31x55ZWuUddfaFQbQ45GEg1cmzPPPNMVT46z/7QwA8+h8DB/Z4PHbYHy0EMP9f3VC9NhptrRGXFCA45GJPdRFGUzPE1aCA3gbHB6mhEX2TCGYKZpbjfddNN+Qs34zxQ2DW8a7qa8KE4WLVr0muFpYDg0gLNhixRe/0xptDHG+Lx589xfG15rCthc13L+/PmusPMaybymPJYP9vWnqDAC2cD8ZBMAmcSNgcJl5MiRfVv7099r5IdJf77l7gfPy1dU+mHEmfc6ZIJpNuXK+q8oivJGoQJHUZQ3jOrqateQozCgwUpD2wglY9xRPGWCRimNJXM+4S+36Ri2F3OMhp399JzGF8WCwRiNFFQmTRx9sA1jG6aX6Tfhc4Qh0/s1fLJPseU1qv2gUDJ584PlY+fD+M8UNhd3IPZIGY1474gIDVAz2uLFlDHFXzZYxqY8jOM14X6/9DEvLMNcoxeEIsA2kimWc43emLpG5xXTfhysAMh0DQ7kGhGKv/6kPx/yFZWZYD5yiXliXx/WK6b/YEd9FUVR8kUFjqIo/YbGizG6aPQaIyyXIUbhQAFBI4kjJLmeivthRIR3pIfOHoGgMWlGKZgmHs+GMfjsEQFOV8sEwzeiyIz4ZDJgOYpxoFODDEyfXc50B2qkUljwGtxyyy3uthFgRth5MeXHMs+Gua5m1M2cl6lcCMuQYss7Zc4LRz8YnjGSeW1zjd6wPjB+OqYtHw5WABxKWI79TX8u+iMqDwZeHzNixXrF9B/I1DpFUZQDQQWOoij9xh5l8QqNXIaYGX3wGlg0gnluLkOX/hhnPu8C0LCikcupZl74dJnvXRi86WHYjCPbKAHTQiOUU+uywZEEu8yywXd5aMh788a4WL7MDw1U/s19xn8mzHQ2r3hhvsyIV74CzG+amR+8xkxnPka5KcNco0P0x3wzrSwbTi/MJMj88KsDfhwOAXCg18gm3/TnQ76i0g/mg9ciG7w+DJ/X34hxnpdpNFRRFOVQowJHUZQDwogZTjvrLzR0aFx7pwJ5RUcmaIRmEi5ezCiLd6laGuGMzxhgNhwlYN5o6NrTwfzgcfrLtCABjUimNV/MU28zuuKFL4vbGP+ZplVxNMIvfnMe05evAKNha8qTRqy33Gzoj2WYzxLBLEP6zSVyKMKYVr4Dkk14+sH0GMOc4iXbNDQjAIwAPFgO9BrZ2Ok/WBhOPqLSi0l/pqmMBl4fe/SJjqOcFDk6TU1RlDcCFTiKohww/X2KbqCBRYOHhp1t9BkxkY9RTIOPoiKbkW1gOmmw+okcY4DZ0+WYJ+++bBh/fiKHU3UWLlzYt5Ublg1HqGhce8OjccgyszH+vWVJWDYUDiwrLzyPRjWN3P4IMGIEII3YbJhyz2c0xLvogR9GhDGvRjT0F9YBlm024WoEwKHiQK/R4cSIynyhEGb6WS657nm/BTV4DuPTaWqKorwhSCeuKIpyyBAjZr9vZbCZETHTt7U/m/q+p8JfGzsMMahe990SGx4TA939m7/mbz+8afOSLa0kn7TQj83BNLM813YMn2XllwZTlrbLVhaEeaU/b/kT5sMbnteZ85gev/3ExEFn/03nvRbe68frZfyacuVxk39vGk149j4/Z+LwpsdbDn71xXtOrjK2yecaef34XetM+F0Hhp8pLJMXQ65rTv82dth0P/7xj1/7m2Vn4/WrKIpyOAnwH2lsFEVRlEMMn3o/+eSTb/jTeUVRFEU5ntEpaoqiKIcJTrvK9b6CoiiKoiiHFhU4iqIoh4EDWelLURRFUZSDR6eoKYqiKIqiKIoyYNARHEVRFEVRFEVRBgwqcJTX4PKlXLLUdoqiKEcCLtDgbY+8Hz9VFEVRFD9U4Cgu5hsHnLFo3KZNm1yjItMHDA8FDDufb54oinL8wO/s8Ps8bIPsNmnChN4v4yuKoihKNlTgKC7bt293v8Zuw4/T0ajgSlA0KvQL1IqivBHs3LnTFTNsg2zsBy98KKMoiqIofugiA4oLxcs111zjrvqUCRoVd95552tfE1cURTkccCoaBU627omjv9dee637RX5FURRFsdERHMWdInbmmWe+9mS0pqbG/fU6wmkjfscyufLyct/9B+IOZVh+7qMf/aibRy/cz/eTFGUgwYcafvfBoXDFxcW++/NxbH8obgi3hwwZ8jo/dGyvOK2Wf/P+tPPjHd3xe5+nPy5T28B4Mx1TFEVRjiAcwVEUMRb4qNT9zcSSJUtcP/kgBopz55139m31wnMZhs2iRYucm2++uW/r9XjT1Z80MFyGb+NNVz759sbHbeO8+TF4/Zh0G3fddde9zp+dBh43+5lmk07j7Hx5jxn/th/7uNeZtNhx+jlTbgzf3m/gcbOPfnJh+zfOC6+h14/tTNr98PNvO5aRN8+8Tna6vPXH9utXb+nfL+/ea2TDc8x+b3zEW97ea+s97nXZ7i8bk0b+GlgW+VxLA8/33hOm7ufCe65fWDbecJlOppfOW47efPi1DawLdn3yKw8vDNM+nu1a+NXVbP7pzLUzacnk/DD12K8MvfXe63iuIVsaTRna96lfHSZ+fvhr9vk5u8zsdNh5ssOlI/mkORfeMvI7L1f67XK0Mdcmk7Pz7c2f1zGvfv7sMjL3inF33XXXftt+ztRrO63eMrD9M35vvrjP4E0Dw89VD/O9VoriR+5eRzkuMB2o3Vl7MQ1UPrDR9TbuPNdudAkbMLsR9OJNV3/SwHC9DaR3X658Mw9+jSz3MY92R2TgOaaD9cJ93nJhWAzHdFQG+vOGTz9+nYZdriZPJjyTHoMdhvFroF+/dNO/HY9328D02unLBNPgjcfkxVs+Zr/3Gnnz5YcpWy88z8Tjl2fGZa6xNz/c78034TkmX37HCY8zPQzDhnF44zHp8qsD3jj8wiQMk/7zhWHY6cinjG0y5Z1hZCoTg/fcbOVIeIx+DCbtfmn27qM/b3mxDO2yNveGt94ZGL9f2WS6FgzLTq+hP9eO2/b1ITzXL1zup387TzZ+9Z4wfO5n/vzK2A6Px0z5mPKw99nwmF/6/fJJ/NLHuOm8+xm3KcP+pNkPc91NeAZTLt5zM6XfpDUTJh6m18ak3y4ns88btzcMs82y80I/3nRm8uu9TgzX1Cd7P+F+Ow+Z8mXwpiFTPcyUZ0XJF52ipmTEO4Xljjvu6DvSO0fePuZ10ij1+ewfnFdvwsh3WhiniNjnmDA4fcXL1q1b+/7KjyeffNJ9N8kPzv+/9dZb+7b+BRdl4LH+cP3117tl1t+pcEuXLnWn89hfyzeLQxBeJ6Yn07tV9CsdTM4Xti+//HL3l/ERvvfAeO2yYX156KGHcr4TwevFvJo0GpgHpoXTIHMtB8xplXwX7MILL3TjzRczlYnpZLnk4qabbnLrUT7LE999992uf5aLfa944bV+8MEHs5Y542M5iDGBn/zkJ317e+G1ZBx8ET8TJp+8Fv0po3PPPTfrPXKg07F4T3rTa+5ZunzCZR7sc7xl/MADD/T7/j4YeC/05z5nfed1y7VqZH+v3Y033uj+2vXJ1NdMbVQ27HudZZytLRcj1K33BqaXebT3EeaBx/KB+aDj/c36771Prr76ajeOTHWmv2n2wnSK0e7WJxteDxEsOfPBvPIa8r5lO9Nf2A4yf4sXL+7b83p4fZl/tt/Mj7kXuM20+7VtbJvosmHqJtuY2267zf3bpj/t4YFiys/0B9mulaJkQwWO4gsbGb6Xw8aTHTOd3VgbI5rHie2Pjg10f2GDTUPIhEFydSbs/NiB0z87NTa+PIfb7BwNbJBpFNGvMQjygf4zLapgDAE/w4LvEfQXNub97TyYBuabefMaAuykeJ28HbUX5m/evHl9W/6YTmb+/PnuL2EnyrjZGRF2fvmIBpYpDQU/TFnfcsst7q8ftsFHI8IWd7kwRkM+5UKYHhoM+RjgNAjoP5dRybhZNylgMmHKO5NYZDlTfGfCNo5ylRHrjiGbUUWYL1P+rG9GbBhnwzIzfr0Cm35ZBva9no1c7ZHBKwYPJ7zepg3IF94jFLfZ7vH+XLtMsP7w4YNfG5UL+15n/KyvmWDavGXA+u81jmmAU5jkg12vWf/92l/Wf7su2hxImg0Mj21apodazAOPZytP8xCIZHqwlA2TBj5syMSyZcv6/urNj13vmXZvHePf3OdXlgb6sdPrl/b+tIcHil1+jC9TG6gouVCBo/jCRoYNmd25+hkUZjnXA+mEvbCRto1ONmy5DGYub22MZXZqTPPChQvdbRvT4dFAyjet7MQyGeIGHrfTaAyLA+FAOg/mi8Yer4G9AIRfx58Nr0FgwjGOwstbdvybRiqPUeSwQ85VtiZdftfIwLz4GYDcz7TQ0O0PNIRMPtjJ9xfWy1wjLsyXEeP5GJWs28xPpmtNI5fHM8Fyto0awjT2N58sZ9Y5+zyOMGWC9d08LWZ9NQ84aJD5CRVeK4bL43b9ZN5sw4V5yZbffNujA4FlwDRlE6VeeL35MCabIe1HdXW1+2sbqORArp3BtDe28WrEtmkTs7WjJl7j/O71bHjLgPWf19tug3itsoVH/yb+fK4Dw2JdPNC2NtN1M8Z1pgc+I0eOdH/Z79jY6Wf59QdzjxjHbbarfoa9aQOzPRwx9cAIVcK/M/Vl5r7Mdv/Z5NMeHggm//0tP0XJhAocJSPeTqC2trbvr39hOuxs0Phlw0W8jXm2zpwNaD7GrG0MUyDZHb0NOw1vh5itUadRkOupI48zDyYNNCxyjYZk40A7Dz5to4FpDE5TzgcKw+KIEmGY3PYzUIyhznz7dciHEmNEmzzmCzt2nkdHY6+/8D7INeJiP6HOx6gkPJ7pKfSBwDj7m0+m1Qglc50zGX+E4odpNvXdGJp+I25GtNjlT5fp+vH6ZiOf9oiig9fJNjjpsl07c337A6/3gRrXfvTn2tEAtPNm7guDLbYJ02m3UV54bj73er6Y+m+EMNuyXCPxrCcm/5kMcS+sX8y7GUE+ktjp729dYpmbus9zGUamttRca3O9MsFwbKHA9jnTgwuGxTBNGnJh7pds99SBcKDlpyiZUIGj+MKncF7jy69D5xMtNoyZOk9iTwcwHahxmTpzY6hkMoYMnEaRTSTZsNNgR2sbB5kaU+aHoiFXR8/j7Nz4hOxAn+raHGznwfSYsiUH0/lTKDItFEvZri+npDDf+WDK0/sE1Ib1Kdv0jIMxvrKFmw0j5DKNuPBescU762Suesl80Jjzu6+Yzlz3VTb6k09T5/J5mGD8sizywU/Emetn181c76Xk2x6Zd2Jsg5Mul0HI68t8ZRtZtGFaMj1IyYZ5Dynbg6Fc147ptPNGZ0NhwfSZumjaEu+okU2+93q+8NowDSSfB0U2+V4DwutKQz7b+2j9wUzB9RPPxOy3p+p6yXYsE7yv2NcxL/k83Mr1EM0eRc63X+LxfNvxXO3hwXAg5acofqjAUXxhY0cDyzba/Bo/+qORlq2hM2ERM8SfCTbIjMu8R5PLvzEScxlIBj71sw0DW3zZULAYYykXZs75oXqqazqPfEQOyz1Th0jxeLAvXTMtDCfbE9j+xsHrlWkagjF6M10XQz7vz/jB/ByoQDLGule4sPyZJ7temfqey1gxT6G95WHyn+1dJN4nmbDzyTLNZYiY65yPgUO/NIZNu8DpR2YUyAvTQCPUazjz4QHvGRNGLvJtj5g2xpevoWbDc/MRLeZ6Hwh8/4/3drY62N9r54V11K6LdLy2uUYUTR3INdqSD6Ycmf58HhTZ8FxzPss6W9tOf0zzoRpNYDp5fTK9jM/9ua4fj7Es+wvP433FvOQSmbwfzD1Hv957iMfNNe9Pv2S3q7yH7AcKXjK1h4yb+IlOI7ayYZdfPv4VJRMqcJSMsKGyO8lML+ezoWUjZz+R9WLCMo1fJti480lWJoPJD+OXjfyhePpIaHzl+ySJxiiNLzb2+RhI+ZDLGDHwaSfLzGsE0TDgNcn2PkW+mE4vXxGZC14vGgneTpn1h4Y+DdRc9eRIwI7Xz7DltfI+oWb68zEqid/oAs/nftYpb7nz2rLsjIjKBusFy9TUAxoMPNdPeJnrnI9BQQPEtAt+LyPb8J6g4cZrbhtMZlolXT6C1bQhxmVqjxjfgQrgfOA17c8og4HXkdcs33eHvNcuHzKJLxq4bA9ytY+H8l43DzLyfVDkhe0B27Zci8L0p6/IB14flpW3DLjN/Yfq3S8/eF+x3Hiv5APvJ/rl/eWF5cb0Hki/xDaC7UAuIZdJ6HM/r519v7Pu8QFFrutpYF2m/0xiU1FyIh2Fory2dj1/MyHCw/Xjh/eYNLq+6+t7ESPwdevq23jTlS0NjM8+xnAZfjb88s2/mf5M0L9xJo+MRxp192/+2n6YZpNu44xfe5+3vOjH+DPpNM7ki+HSj8m77bxkCsPgl26Dfa6dTpaTfU5/8MZHZ18HwmtoH/eWUTbs8zJdT788e8vSi/HnLU8bb7off/zx/bbtsiXZ7gP7PDpTJwzea+B19nVmvNznDcPGTgvLItu94IVhe/NmYJi57iv73GxhEZOXfPDmg2nx1n8v5vp662S2OLNdC7+8Z/NPZ9LorWt+aec+c9y+T7znMu8Hcq/b/r15YZjeY+b6MLxM6bfT7OfseOyyMvXTkOl6ZgorH+z4Mp3vTT/zmQ/eNsabHxM3r5NdtplcpngZjjdsYq5NNmeuvTetXrjP7z71nud3rrce+jlFOVAC/EcqkaIofZiRqAOZZqAoRzt8ksqpb4f6yffxAJ8qcxljLTtFUZSjG52ipigeOD0t1zsginKswqlPh2Lq4vGI33RERVEU5ehDBY6iWJiXGo/Gd0AU5WChuOEceK3f/YcjX/19YV5RFEU5MugUNUVRFEVRFEVRBgw6gqMoiqJkhCsqGXeoVilUFEVRlMOJChxFUY44Zulj4/yWMVbeeLg07s03935Ykr/ZPhZ5uGFa7DpyKJYyPt6gQLXLkE7vNUVRBiIqcBRFOeKY7x/RkN60aZO7rYbXkYff0TDfg+KqgofqO0/9xXznifXDOH7bhQa6Oaa8HopA+7tGEyZMcL+vZJcjF05gOdrfLFEURTnWUYGjKMoRxUx7GjlypPtrPnKZz0cylcPLokWLfL9I/kbDOnLuuef2bfVCsUUDfezYsa6BrmTH3Gfz5s1zfw38uKcRjCoWFUUZKKjAURTliEJBc91117nfZjHQeOXogXJkodHL0bQj/e4Nxc3WrVv7tvaHI0sc/aPIOdLpPNqgeOHKb4T3GUdwamtr3W0vxp9O/VMUZSCgq6gpygDBfopNwWA+RsipJ2eeeab794FQVlaG1tbWvq3Dz+jRo7Fjx46+rfzjr6qqQkNDQ9/W/mVgwyk7nAZ3PHO4r2m+4VOYZFp2mfX2jjvuwK233tq3Z+CRTzl56/Ubhfc+zJc3ur3IBEcfKdqMcFMU5fhCR3AUZYBAo53PK+geeuih173DYo7RqLS36SgG6Ox9Zr5+S0vLfvuN40vnNCL8jhnH4945/3TEvHNDZwQHf7dv377ftonfL910hMfq6+v3Oy/T1+Zp8PA8ps28QG/OM2limpl/E0d/nAnL71gmlylvtrPT6+fyuR7ea+oXrzccv7qRyfU3/GzflOHoEa8h/eVzPfzS6VdmB5K/fM7x+snk7Dqa6d6iI3712hzPJz5vOv3Kw+sI4+V96N1n+/OGlau9oPOL35xn7zuQdBtn0srRK05z1Wl3inJ8ogJHUQYI9scbL7zwQtdAyRd+nf2NfFJO42Xp0qV9W0cGCh3zAj3LjgbUTTfd5G4fL1Bg0MA8XAs6HGj49M86rCgHA+vfQB4BVBQlMypwFGWAwak97NSN8Z4P+bxITqOT0+Do7JWZDHxSao4bl+k9Gr4YnumdijcK5oFToAw1NTV9f/ljlik2eTd5/Na3vvXa33R33323e9wuL7+nyPY5RuzlOseG19kOw2Bfh3yEhRkd6A92vHaa/cI6kPD55J11JBve+nY0YZeJ11HwHS4OpExs/7nq3NEC37Wy022Xt5fDKeAVRTl6UYGjKAMIdvB834ajEWb6T66REhrsfJGc52SCxrS9lDMN0BtuuMEdBTGGBUWVd7oIp5YYaDxRDJAjLW4Ip7BwKh/TToPpySef3G8UzMakm3niyALPMVNmFi9evN90mttuu831y/Iy+7xTBlnm/T3HhunldTZTeygieD0oKHnM7Mu2QIARZjQA+7P8M4WeSTvLwK4XtvF+oOET1qtsgpN1ieXDOOkY/6F4Uu8VjcaxbPuDWeHN3FMmnXRewWeEs9f1lwMpE8Zj6jHdocIr+o07VAuH8B406WZdZB1kufrlgfdaf0azFUUZGKjAUZRjHBoOxog1hgoFhDEqchlnNABoLHA1qkxQJFGsGNHEX55Do4JGNv+moZGN66+/3k0L00TDi86k8XA+1c6GeR+HK0sxPUyjHyxPs0wxp/ORyy+/3C13Gm32eTQyCQ1Mw7XXXvvastc8h+Vmn2OOec+h6PKDH9y0hYOZYkco3Aj38Zpl+jgnBQTDYBn0B/q3y4CYemFzoOETrwjwwjK2pxOa+niwmKXKibmX6A40bNabXPeFEdWMw46zv/S3TCjmiH3f852nQ3EvGnHK+mfnyX7gcTCwfpi6x7rINGd6OEGOhgcqiqK8sajAUZRjHBpQ9hLLxHT6NCpso9mPG2+80RUe9pN+PlXOZWTSaKDBYozsTCMFBhogTKtJl+1yxcX0mJXgjCgy7kDwTuVh2CynfIwk+3s9fv6NoLCNfp6bLWwzRdB7Tn8w18Mmm7gw182MThGOGOUzWmHSyzzx+vlxMOHnQ3V1dd9fvdjlZUZiKD6NqDYuW/zMTz5iZuHCha8JWcL4/MLNNc2OZFoMIxsmH1yIwSZbmfjhJ2bse9GMxBDeI3Y5ZhuN4bXPJWZYFxgO4+MIjB12PqNxZrlr3jMHIqIVRRnYqMBRlGMcGkg0tmxDku+WmHdFckEDwaxSZQwMGnq28cOnvDRozFNf/tpGCN/3yWf6iTHm+jvX3zbkvOIoG97pP6aMaKDaUHjZ4sKLbdB6jXqea+eHxpoNDTmWlRn14Pk0/mxRmuscLzQgaRjSH/FeD8I00Y8RGn7QMLQFANORy8Dn9KBMI0teDiT8fODoll0nvPnntTT11zsykit+1vVc9YplyjSYfNH49zPoOZJhC6FMML5sI6jmfjR1xuTFvra5ysSLqe92u+G9L80oCeEDABMvXS4Bw5FEM5roh512v1XUssG4j/QiJYqiHN2owFGUAQANSb4LYgwuGlXZjAsvNDZsA8PvqTKNZfMU12vQGYMyH+HCtNHwMsZ5PjB9RoT1B5aBnS9jRHrzyzRRDGXCGGN++TNlZcreNs5oPBqj3hZQTBdHvMw5dllmOscL001/fteDeWEZ5xq9I3Y50OWC+e3P9etv+PnA65hLYPCeYH3J9i7PwWCEkHFmCqONqTe2iDhQzH3D/JoHDTb5lIkXlhHf/zLneEdhKazMiI49fe9QYeqDd+QpF6yDFM5+5aAoikJU4CjKAME25mm4HGrMyIVxXoPOGGC5jDmGQ8Obxnl/DD8ai4cjX4Th0mUTaCZ/fkLILnv73RpjBBthZWOfw2mChmzneDHn09nXw4SdTSAdDGZaUTZReDDkI2RNORnHeuWF19SIjCOFWZDjYI1xc//x3jEPGryCJJ8y8cIyMv5Zb7yYePMJ60Bg2P2tp0yLKYd8HqrkM1VQUZSBhQocRVEOCcbooDGXCxo0xm+ud3feKGjoUcRkMkSNocepeP0ZfRqImLIg+RiY/aW/33E6mmFd52jcofrGEsNj2ZsRzeN1FMOUA+/ZbA9KeF8frlE8RVGOXgLSQByaOQOKogwIaDDxyaiBQqS/T1iPZWiwc6SG0ICkgaS8sVBA8j0fv6mSyr/gCBpH7vIZ7Tte4UiXmjmKcvyhIziKorijKGa6kXkySsenxN6VmgYifAJsRmVoVJv8c6GGQ/H+hNI/OK3saBnZO5qh+O7PB32PN/iwJteCBYqiDExU4CiK8hp82mlzvHwgj0Yi3ynxTrfS0ZsjB99LOhzT3wYSrJ/H0+hqf+HDGR0FVJTjE52ipijKa3AUg4a+zfHURHAUy17umk9/1UBSFEVRlGMLFTiKoiiKoiiKogwYdIqaoiiKoiiKoigDBhU4iqIoiqIoiqIMGFTgKIqiKIryhsIVznQRicOLlrFyPKMCRxkw8AVxrgJmnF/DzmV/bT9exw6BL9r7HTPOLKfsDcsLlxe2j5uliM1yxN7jXsfw8yWfvNvH/VymZXlzpZPnsdzsfXb8/Ns+Rr9e/34u32WCc5Wzjd+19WKHZ661Fz8/9jXgcT+8fphHs01nx+c9xvrAfd40ecuSx7nPXINscRi8ddkv/fZx+9rY19fUWW94XmeH73fcdvnS32vr5/zuGz+8+bOvS67yto/ZddSbNm7TGT9++bMdr7mX/l5XvzBIJj+50pyNTCucmfrkhx2XqWssX3u/12XKkx/ec4nf/UKypTNX/bfvHxuvP/s+9sPr31vuZoW9/pSBogwYuMiAohzLLFmyhAtlONddd13fnl647VfFuX/RokV9W/+C+26++Wb3702bNrnnMmwbE5chU1j0442bYXOfnU4THuOzyRS/l/7mnWn1+iUTJkxw7rzzzr6t/bGP8dcOl2GZMiM85heO9zxDJv+M0w43EzzfG65fOROG6fVrys+bBu43/r3XhvCYXxq5bY55YTgmTO919YZl0mX7M3XCDtuUq51Gc663XnrjIOZ8736e65cHU3+8xxiOt7y57XdvmDKyyVYv/dLhhX6YDxtTDn7X1ltmhP5yxWXKK1Pdsq8X93nL1cD82vH7lSnPZZgmDJ5j/vbmgfvtNB3IdeV+v2tgyoXh2eST5kwwTL+4iImLfvzwqyuMzy9f9GeXUyYypZv7Mp2fK52Mm2n1wn08z8ZcL2++TBze+E16vXHTvzfObGWtKAMZHcFRjnn41X1pwF/3NJDb0thnfAJn4FMvugceeAC33XZb397Xw3D4REw6l6xPKOlPOhr2YH17euHXxnlupqd3hMf4xG78+PEQIwZ33HFH3xF/DjbvfNrKp3sPPfQQbr/99r69/4LHrr32WvfDi34wnsWLF/dtHTwmvfy+R7ZrQfpTzixTMRJe55fXUwwAd2ls73W58MIL3fDvvvvuvj29sEx4LBMsL8blfWrKcHgsH5YuXerGbX/jhHXCpN+klV/753XmMQPPoT97uWs/GAbzzbLyfgmf9wLxe3p8/fXXu/nzGw3IhqlrjIvl5y0fP0xZeq+NzYFcWy+sS6zj2dJll5f3fmN95fXauXNn357+wfvPWzdYTmwDeH8xbl5j73UycL/Jo53O/lxXxn/rrbf2bf0Ltgt+9TZXmrPBMMeOHdu39S/YrjJM1mm/9qi/XH311e7vsmXL3F8/eL1vuOEGt654y8vUqdraWvfXcDDpNB9ONvXMXC+/dtzUKzv9dnq97TL98763783q6mrf66ooAx0VOMoxjREapiPzws6EDX42Y4pGooEdhB/shMwxdoKZDH76Y3w33XRT35794bmZ4iB2R0YDzdvh2RyKvBtDhMaTMX5smAZvp+8lWxr7g13GJFs59bec2cHTgPDDXMtbbrnF/bWhEeMVWhSdmcqc1NTUuIaPV5wynMsvv7xvKzv0R6Pdb9oJ82UEDdPBcqA/73X2GvxejHDLlCbmneXGsvZC44pGlt+xTNhGL+uMLd4ywTKjgWcLOC8Hem0NdrllS5cpr0z3Aw1duy3pDyxrlievow3TwvuS+c91n9Efr8eBXlfj365vxg/rtJdcac4G67BfmCxDpsO0Xf2pX36Ye3DevHnurx/0wzqWqU1nXecDB5uDSacRRKaemeuV6fqyjbM/uGzylCm9vBd4XQwjR450fw+2LBXlWEMFjnJMYwyKTEYJn14R75NVYxTSZXu6xRES+mEHmA9GoJh4/fAz3Bk+4+GTvHw50LwzvybvLIeDxWt8Mg8mfOOy5cv4z7eMSX/K2RiwCxcudH/9YNx+BgCNGO9oDJ9cZypzAw0fu17x/GyjPl5YphQoTJddnl4Rw3QYIWPqKl0+xgwFB8PPJB6MAep9ek1oXFHE+Y0E2Nj3Wa66ZtdL44jf/WI4mGvL/YyD5ZYPprwywWtxoGKfookjH8Tk3bxj0h94LQ/mutI4tkckaHyzLvtxMGnmPeXFXCOm2xjvxvjPB/NAwHasU9yfqSwI7+ds6WZabFHbn3Ta9d847rMfPuSqV974mV7ee5kwI2PeOu93vRVlIKMCRzkuYQfBToYu09Nfwg6cfvw65EMJw2c8fFp4uGF+Td6zdZQHCvNgwjcuW76M/8NdxgcCjRh7NIZPt/MRKsbwMU/DeX42IzwTNO5ZNsaQNCLGi7ecaTAdiIHcH2jM01jzjjDZ2PdZrrpm6iX9Mf38O5u4OVjMPWfK9khjxCody4Lp47XOJSIPJWZE0BjHHEHLNvpxKNPsFVMML9c0VRtTZzg1j5jrm03cHAj9Saep//RDTFnlA6+BEUXGeUWLoiiZUYGjHNMYozFTw29GL7J10vkYnuwk8zEYTTzeKQ35ki2dXg5F3s8999y+v448+ZYx6U85m9EWe5qHFxpDmcrCHo3h0+1s09NsaNSYp+E834ieA8E2JEmm918YB/0wP3TZxAfzSz+ZMOVlys8LrxeNyXxHHfOta2Z6U673x8jBXluSKX9eTHm9EUYmxSOvIwUr605/4jyY68p9FAo04jk6xvsxX4FwMGkmFAn2aKUZfelvOBzt4L2XbVTEhg8sco0u2hxIOlk2FDt+7VumesVyZ3kyL0YocR/Tm034b9261c17vtdNUQYqKnCUYxpjNGaaY3/jjTe6nUO2xp5h5GN85ppXThgP47PnQHvh081MhifPN9Nc2OGxA83Eocg7jYF8Dbw3AruMaQx4p2QZ+lvONBIy+TViIdM7C6ac6Y+GRb7lZZ6G8zzzBDdfstUR5ptGDKEI8Csjlg+NnGyGv8lvJrHE8sqVbtYfxpOPyLHrGuPM9pSf03BM2eXiYK6tIZ9724SR7X0e+35l/TXXyQvzZt+Xme5zI+T7M73oYK+ree+Mo46ZpqeRQ5lmI6ZoxNuO9GeamoFtKOtlPiLZPLDIVF5sh40wOZh0so5RyHjrfa565V2MgellOH7tA9NKwcVrqCjHPXJzKsoxzZK+JVOl0+7b04t0cO5+L/QnRmLflj/SgbjnMuxsZAqL59IxHBvuu9laitSk3evP7Lf9+tHfvDOtXr/94c6+5UwzwWP04yXTeZn8c3+ua0Tojy5XORO/MqEf7vOmgeVqlxP/9obJ8LxxcNsOy8Rp6lGmeuUNy5SX91qZ/Sa/Jv1+6eB+G7/0mvC8+7mP/r3wmmSqq960ctvvGpqytMPxq5cmXO+18cOUs40pA+/5JlxvPvLBlJc3X97rQjKVrd91YHj0660X3O9XhrnykClu7st1XflLf3QGE55Nf9NswzQwTAOvvb1t4H5vehm2t64wn375Yvq8fv0wdcUvXDvf/Ukn93nLwZSt97qY8vWGYa6zNxyTLjstJuxMcZrrqyjHC/u3WIpyDMNG3HZ+nSw7ENuPH6azMc7bGRlyhcUOzj7u9WM6qWwu307Je55f3u3jfsZALrz5sQ0b0xEbR78Gv/O8/v2cnyHhR65ytvHz6y1j+7qYcjLppV9jMBhnypq/Zp/JP8MyYfjVq0xhMT6G4T2HzobHGY63LtnXN1McBu9xOpN+G/u499rQv32O997wOjsN9n5vvbTzlQvGb4dFx7zZeMsp3zrmxQ6Dzq+8iF/Z8tp64fncb9chOr9w883DwVxXpsP45a/tx6S/P2n2Qj/MB7HPta8X02IfY1z2tqkr3vjtMOxzTLoz4VdeJo3E3p8rnd76b2P7t8Mh9jl0zBv98NeLtzzo7PQajD9FOd4I8B+p/IqiKIqiKIcdTq/iKpBmOq5y+NCyVo5X9B0cRVEURVHeMPheW7aX8pVDB8VNvgujKMpAQgWOoiiKoihvKFwEJdtCE8rBYxYgOZoWklGUNwoVOIqiKIqiZMSs6GjcoYBGN5e6N2Gq2Dn0sIx1appyvKICR1EOMZzzbBsDmZb7PVC4nKkdfqallJVDC5ectcs9nyVoBzreun4sYKdXr2F+cElrvq5Ld+eddx6ScmO7xeXFN23q/SCnTllTFOVQogJHUQ4x7LSXLFnidtrsvLl9qEQOjQB+x8IYBYyHX7dXkXN4MU+XjZFHx2+E0Eg+np88s27T4DV1keVxNBuqvFbXXXfda9eQqMjpH3x/hoLnYNs0foh4gvVByny+RaQoipIvKnAU5RBijLuRI0e6v+y8aQCaL9ofLObjecYo4BSEm2++2f0on3L44HX1fg2fhh6NZH6Ij4b98Yap6+bDjqyLFDn80vrRCtNsfziRRjU/unk0i7KjEU57yufjrtmorq52H9QoiqIcDlTgKMohhMKDT4jtr1LTEKYRdSigEblo0aL9vrrNL2Hzy+/K4YPiJtNX6fmF/mNh9OJQw7rOJ/D2F+tZPylwMn0V/kjDF9s5AmrD+3XZsmV9W0q+8MHKwYxe+rVliqIohwxHUZTXwQ/R8fY4WFdVVbXfdllZ2X7bfs57TjZXWVnpuz+X86bD7wNxxPZzJBw/cmc+muf3sTvjL58yy6esssWh7si7TPXUhn78zj0aXT7twRvlvGUrot3X34G6fO5RxqkoinIoUIGjKP2ERnAuQ8vurGmkc5u/2fDr4P3ONYaHwbtN+LVs86Vvgzcsv/MywbAYJjHnmXBYFn7CIB8YljedNnbYdpwG7mMYmcg3j7nC8StPLxTFdDZ++7xh5VN+fvH7nefNbz7pNnjTyvPMNTd4675325AtXr90+5UTyff6ZcIvD964MuXBxi99fuf55c3LgdyHfvmwscPLhF/aDqZ8861bDD/TvdXf+I1/47zhmrI1LtO1YLptf/nC8Ozz/Mqc9cT241dH7OO2y1V3FEXJH52ipihHMZwGJJ3eEZ9CIx35fu9aSCe+3zS8w0VNTQ02btzYt/Wvd5CUoxtOy2Sd8Vv8glMq852yybomxughX4lQOfbgYhBcUIX1SmwX17F+GFjXWFdEeL12nHgXkeBUUk6jNH5EcOT1Dt3EiRPdX3Me42F89rRUxsWlr40fEWDulEh7Gh6nuprjtmNYutCCohw6VOAoyiGAHaRxh3pVLYqc7du3920dGdj53n333X1b2O9FbT9oDLAs2OHTADBl413ieunSpa5/e79tDNBQvvbaa/N+t4VGjh3+4Vx8gdfZjutYgdfEpNle5vnWW2/t83FooDDnSllezAIZ+V5TGqNHuv4fDN56ks/DCvsa+b2jYu4v+zjDNfu8Rv2BYt+7tvML3+vHD+/9ScGSD8wjH3RQCJj64+Wmm25yH7ywzTBwMQSKaSO0GQ7bMvvbMHyHjvv8ytnA+4TCyhYgjId13DzoYRx8V8+OnwKMAuq2227r2+MPy1Pfo1SUQ4sKHEXJADs8uzM2zvv0mfvYiZkncbmwjRfb2dCPeWqdryF4OKFxwSeRJp2LFy/OKHJo0JknpITGg3mqyo7eLiuz5DX32/uMQUJogGQyamx4Dg0mPjU1YXmNhnzKPh+YR4Zt4mGch0IgZDIoD3bFKoNtKDLNDNdcDxqHBwPL3zZ87fvE5Mv8nQ0eN0/LyRtp+NmCz3YHem1ZT8y5pq5woYNs8Bxzjej4xJ/3CM8z6WH5mOO8Fwmvpan7hOEcLLzvGB7DJeY+9htpoLFPeJz+mE7vyBsfOLCe0Y/tcsE2gQLGvn/tOkJY3zh6YsP0M13mQQrLym+VPz5EMeXox5NPPvla/my4VLxdPzninAvvhzd5T1IY5dPGKYqSPypwFCUDprNix2Z3xnZHZwxxGuEGdmA06jNhOjLb0KezoTFDg4UdOTtuIy7o8n3qeagx6aTRwjT9//bOHkeyYonC87aBEEKwB4yRcHDYAA6zACT8kdgADhI2BhYWCIkFgIODhMEekBBCbIPXX02fN4foyLx5b1X19PQ7n5TdVXnzZkZG/kVk3qryOjsYqDI0MACAXU10haHm98lwciMBA+jIyQtGDG3jj61U43hF9yuQLwaX0E7tubiR44/iYDBeAoxljDnQ419Pnz49vT8XvhqdfqF+St7azaZetKvigf9K2zkQulb7zDXhq4sB2bx/HHX+qBd9nPaT87flsNW+pflE/QHcSZKzQRnq+9XwPhc5Jn46UUEOzY2kQ9bqzLEpsuXgVjTHkhdBbYKDIydHear9HPqevgERx7HbmKmPwlbI38em8K+6Rvedfti80ZirkC9z+3317xD+n4iDE8IAFqtu164ig83RotdRd/A6WCwpm6AFXUFOwQgWfZwjZJCRSOjkXKGeeijvGXqkCANA5eq3gRycEq678cCi70aITtIA585lqVQjpH5mZ0X3q1Rjyr9GmjogHwYuwWVGfzNWnBnywTByh4Kw4vy6nBh1Moor7HSzcy1oB8obge4lO/20GozPnz8//SdN7dPuQJCPfjfKjfo96CSG+7VRoNA5UwJdbDkzGNzkUzceCLzvkOGLDo9APaTbOo5l3OszcsD48xMOyXakv4jqGHjfEH6ywwlSnScw5OkXLgNBTswMnD7vq5RF/vWU6CGhU7SRA8PjbZfYGAkh3CUOTggTWER90a5owXXDZeXREIy2rV07ysUg2WsUabcQQ8iNyGpsVDCI3OiQ4VB/4JIFudvNFKTX4x7oR4Yu9+CwST/avXRkRPNBdOGv9QiOgkM6jFc3lnSC5HDfuTum6Njzpkw3nFVXwGB2mbccGO1+z3QsI7w6wFvOLydrqzv7OINy1Ai0leo0QpsC/qiaoD7UHYdDRvkIORrd40Qr+E56dai2HBjqPdvNRzY5GfUksBqrvFc/YTyjw+qgVDDk/eSjzie171Ud0Rcpx9Mc7S+CfuOfI0GmmaNIf6GfUUYF3boMtE83Tivd6Yx0qbHSfeaLvibnjDnOHXxRHcIK+Xd9lvJG7Ul+3DPqS1xDhz6/hRAuRxycEM6EBYzFX4bglvG2BwwQjJWVHU6B8Y5hNVuwO6rhISOR/DweeWZOHOlHjhlGHoYPesIwcGOTPMkbg8eNe15TLnSnQIJ0OHF+yrNXB6tQRwxNlUOZ1finrsR1j8ScC0Y4unM9rUCbopNZ+znUwdt+BTnmXRmUv+q46LStc5ZWkLydYXwujBX679ZnLugnOCzqJ7DlYKEj71uMFzeiyRP96brrkk0J+iJluJN3tL+IKhOGeeeo6+SMayPDnnlhzxypTSR9jsZhvKt9GWv1VIly6It6DJPNl87BZ/6uGzkODh75VDhp7BxMdEC7MxZG8KUtyHy0TUIIc+LghHAB3DmYLWp7YXHHWMBo2fMoBkbQJeVwqCMGzuxkSY5ZZ+S6rp49e3Yb+8III84NM4drW8YA15U3oX4GYBWcgC3DGh17WZ1stAHprgH6khOwB2SSk3ktMDzpI137IzP6XSmfvoJh6W2B0cq9KyebtIsM5EuDbKO+6pDG+4ke1ZvhfatzFDRWCN4HVFbXL4jr4p2Zbmt/r2DY6/HLkV7In3kBh01lcA/Ok5CT5Js6bNhwnztG9AnyUfsy1ulzPk/ifOBEKA11oG96v+Q1cbNxSn0oy/sh8tM3vT25zlyNflwH3TjglDvOTQjXIw5OCA8cFkqcnK3PbtwnLODV4HAwKGTk7jl9ekh0O7aPBYzmI6d8q8jRHLU/ThaGZ2f4VcgHeauDP/vWq3AeR3TLmNfp1GjM0y/0GB/zB44Mc1vnXPiJDde5DyeDewjgzp/mHP/MFQ5O3eihP8nBJ/CauC1Ulu5Dfu6TkyKHB5RGoYP86re+hRAux39uBujacwchhFDAQMaoAIybrR3i1wl2Y691ChbOJ+1zPe5DtzisOCOdCYKDhIMzO1UJIYQZOcEJIewC40cnN+xCYqAQXufTmgrG1+yZ/PBq4eTn6OOHYc596ZaTaU5lOOGoJ8F86UCcmxDCOcTBCSHsAsOfR0X8kaHRo2qvIzhpfFg5BtbDhEeBeLRHn6sIl+O+dcsY4/N6/kF9NlC6LwIIIYQ95BG1EMIh6rPlPEs/+nBxCCGEEMJ9EQcnhBBCCCGE8GjII2ohhBBCCCGER0McnBBCCCGEEMKjIQ5OCCGEEEII4dEQByeEEEIIIYTwaIiDcyH4aln9anH9dXDe6xq/McBX6ur9KOj3RDyOfOq9fKUm1PiV3yPx9AQY5ad44df864KpX71GnOTha0j9eg3d1w3XNF4eVFm7PGZptmTuqPd4Wo/vfim+6qBLg749zUhebxdk8Hu4RhzphV/vQtXtHqRj6lfp9EVZHleDy71FzZ+8kaNrw5mcW/2zuwdq+bxHhq58qOV0fWCL2kfU1t5XRJWvjmfw/Eb17NLofRckSy2/00tNU4NkVvvVePA8eL2Kt0enG+jSrMoMvg50wXXu8d6eXt5sbun6ALLoXi/LqWmqrrvQlQUa3yttrTQ1HqrehMd3bTbThd87mvNW0jhVV2ofyVavu8ySVYHyVnQvvUnXhKoLT0853KP3klF4nbvxo/apzOq2xVafHvUVx+tE6GTvWJXb0xC4z6nXPXTyhnuCb1ELl+Pdd9/958MPP/zn008/vY15Ae+/++6723cvIO0XX3xx++4l9X7S1fy6uN9//51vxDv9n0GZpKtlE1fvH8kIyOlpeU9wkJE8VXfy02v+c02Q1suSnJ3eajlA/Ehe4nTdWZF5RJXfIc/aPkD6KoPy+eWXX25jXiCZO7jH66k8vD3Ij7hav5FsxNW0exjpWHT6Up+tdZfsXVs6ur+WqTbs7t+SU2W7LkHy898hjuBQxqh84mvZyrvqoWPWrsS7fNJPTSv5ah1nuiEt8Z2cXZ9SG6iMrv0rozRdmyj/KsuoXbcgH9Wv6gWkl9qme2Xuxpj07kinNZ7yqq7FSEZxtH1HedY2dpCf+0ayjvTWxXV6Q8ZOl0L1HOlCuu3y2OoLFemBOgnuI67WZyRT119glF79Q3CfdF7TE9+1ac0DeO/1cKST2fWRvivST6d/ofoI6cjron7kcdxT61VRXl4XynN5JKOjtlY7kU+nD/Qw0lO4H3KCcwX4Fegjv+qu3YMff/zxXz909sknn5zyc7q433777cnN4Hzyzjvv3MbcBZk+++yz02+W1B8yvOkPp/9///336f9efvrpp9MvUDtfffXVqSx+OJGykXv0Wymk/fnnn0+vXc6anl/Pp6xuF5L8v/7669t3LyGOa5UtmS8J7Xsz8Z7kd/SL3u+///5tzH6QtbY9P9ZHm1LHGex2sSNFvaHuTq2Cjr/55psnNxP/2btWyI6u1B9G8AOB1LvqlLrcLERP/vjjj9uYlxyVU/3Q+4XGrMaOYGzRplX+S/QB0lA35gmH9+jizz//vI15qZ+aFvnIw39gUTBOOt388MMP7Rga8fz589N/7jsKMjDO6Q+MSc+LNqZurjN2gM/5kUj0QftUmZGj01XHTOaKdtRpD/KvOkeHtEU311UkI6GbA8Wl2heePXt2+l/XDM0h5FfXqRHaTa9jqYMTi88///xOvxZ7dMH8WOe8b7/9dlkXyEIdbwzdf61VzMXoGY7OqR0jO0GgF9bOlTL32CrUE50w5pg/z4W2Yc6btSFt43VkTDFvUUdBHciHawL5qNdMB+RBXt5mzCmUKX18+eWX/2tDQRqfI3799dcn77333um14H7WHs873D9xcK6ABmE1nLdwQ9QNoI8++uj03ychGW4ex6DeKpOJm8E5GngsyAzYI1Dnjz/++M4xM2UxKaCX6lRVSAfICSM5KYtJvIKu6uLNaybTji2ZLwUTLe07WjTVxiuGTAeGBvl3R+JbBoMvICw2Mwd5hHRMG9O/1H5HIT/a8YMPPriNuYvSsMB10H7VADhHThZ44JfeQW3KmOmgr/s4vkQfkAxyHirU2TcJ0M9oTqDPcF15irfeeutkxFTdYChKxvvC5yLGZJ0/ZBxh9BGYA4/0X4f2qUYxupAxv8WWzI47wJojK/SvFYNVMqpdaf+OS7YvfYtxVOXG+OOa8qt9rELdyGfVuaGdR4YxrOri6dOnp3KrE8q8wbUVkAV9dm1HX8T4ZvPxUozsBEGfQ56Vx7SQedVWoZ60J2mRYas/zqANaRfy88e5fO5jHNE2dTzTptIBMpBPbSvN7zO9k4fmckFZ6M7H8NZ8wviuadh0uaT9EI4RB+dK0LkZeFsGKwuXBvcIBg+D1RckdidqHAO27iRUmLi751zF1oI8gzqzGFNv1Um7TatookBOJpoRb7/99ul/nWQ1QbletNh1XEJm3efBFyHQRDtaNFXv7sRhBSZ0GQdMrpJjtAjRf5SGup+L6xgDkfz34DITeI9hMOuLWoTeeOON0/+OagDskZPx5TLhCNNXGCOgNn3zzTdP/zu8/Ev0AZ0eKW2FeBl+0s9oTpDe/MRHYMS4buQY7kEG1hGnSDrvNjEq7Jwz3nCGOyNzL8jLmHCjmPloK+9VmZFVaes80bFqsEpGGXc+B1aOtK+vVQqU2RnZOEvIrfkY43iEnBvG+xYYxozD0aaG2KOL6tBSxmgTooM2HI1HYA7TnCE6XfpJZGXFTnBY15Bry7GEFVtF6wj1VF1mJ5NbMI/RLvQ51i0CMlBPycE82NkqmreQSSeH3RrAvd3cBqpPN3dTR83B6KbCWEHW0bzGOF3py+H6xMG5Iis7bwwEDfAZTLhM2sCkxa5LjWMhmU209wGTn+qDPFrMtxy9S1IXb17PjJNzZda9HsjnVaDy6XvAItItEvQfpSXNucigAE38K4urwEiVo6UxcdTRnrFHTuRBLkCfyFQNlRE4yTJI7rv/XwLVU7rBQKy7nRXGmdeZcYQOj8xJ6psrhsJff/11+s9cO9qp3wPyMn5lFKODFeduVWbyVtrVeWLLYCXejXJe+xxYOdK+vlYpoBfa2vVenaWtHX/SkvdW+5GHHsWdOQN7dVEdWpyx1dObo3S61FzT4elXoA9zD87gClu2ik7kBOtHPeXcCzp35xiZJcdDRptv3byG/uhr11i7wn7i4FwRFhEm45WjYpgtdj4JM8mzGNU4n4BGsJiwUNwH7CQzIWvx2gNydjuDgh0WDPNukvHFm8BkvMo5Ms/QLrqMsYoWli0jYxV0QD3oHwQZMh1ujByBvClDhq0cJp02rEJbssij95m8QkbI6uMfR+TEGWLRxVCoxpfa1B9nEPQjGSyUyYJ3iT6gazNDUEg/o8/UKX5kzDFutPPOoq1xNcKdZoVufO5hy9DU3EdZW4bvHnyTBB2MToA79hjHs0cwHfQ4M1iRkXGjvq25a9ZP9rZvB3M048hPSHhNfpJFMo/GKXkwPrbaj+s4ejqhHJ2079UFuiVv5GYMIs9sQ6zCvdpovA8obwV0Stus2B9btgrODO0ondK+zGsjh0jMNno0/zp+osJTGp0NoHmLdtPJTTfHcS+PY3ZoXurmYuqkJ0Qc+g91YD0YOTB8ZmePvRGuSxycK6OdNy2WMzRxAxODG3kMSCYEjCktRjVu6/E00EI92lFmcPuOP69Hj8xQL00U9T5HC/7WZOgg58gwJx/q67t0FS3eM+Pk0jLPQE8sIKNHK5gYwY0MJmd00IFR55M3/aVbwNVHRkf1QB9VO7LA7T1tQMcY827YYoyt9PkKhoUMuS3dk5a6zR5Xcb0clZM2oT9V40ttKgOqUh+bONIHKrrmH7Kt0KfRnfQzSks810fGHOOGMU5/uOaizRgfGVbIJmOCdvTxSh1xRPQYiebP0Zjeg/RM3fcavDOZK6RT3pQ1M0ZlsFYnBz0go/drAmnd8ahcsn19rDKeqiz0e8bfjC3HxWEcS3bnqC7k0HJSMVtXOrh3tFYBfWClTqvM7IQKOqde6GqLka2iPlx1CluPqSGr36NxwUZN5yDhcNBWwBrcpaEd6U+g9a1uMiEz985sIvKom1uUhQ7qJgXjkvmfOszmZ3Q3sjfCK+CmwcIFuRlst69ecmNEMRvc+cpA0nKtQvwsn5vF6DbmZdzNYL2N2abLBxTvIDNxVc5OduJIezOx3Ma8gPhaFijvEZLH9UbexHX1dZ3dLICndB5Hfv4e9srszORHvu7+KhOQjnhkrnTpu3pIV7VNFO+MZCPe06p+nVwCvVVZQO3kbdfpS+lqGdLJFrp/lK/K3yOn+g7XHPTT5aHyax1I2+VDXM1n1gcqkq/moXjPQ/Wr40VtXWWjv7guVAflqfyqnKM+5XTtr/4pObo0oHiXrUsn+VwWxdWxUaFOfp/axO9DHzWfPTKTZ20LUFneHqTz96A2djmRp6sbcVUu4lye1fYlritDcitP/rtsQrrw+ijOUflVR53edL/LxetOTuJqWeTp9eQ6QTJK1y6z5KtlSA9dfB2nXR+CrjwYpSfe8+a+Wd+qbVrlAulJ7Qnc7+/Fnrp1kNbl7XTLdS9Dbe516eJ4j3wzpG+v20immhfxtZ6j9guvjruzcjiEOrdC7eQMHA0kDZpZ6CYJ3ecDsotboZNhNDF1aX0yEUwMpKWunrbLV5OuQpcfVL0San5VPuWFHEqLbJ6GexS/KrMzk9/jybdSZSHMqOm7iZv2R2aCp63l+7UueN7kSdxIF15Wlcn1ifydvpS/Qi1HeXT1rVQdeb33yFn1V8eV4qustX4E+hV5qq85VV7CXlx2AnmOqGldP8Jlkp6op9KO2svjunyh04+C5J6lUajpJKeoekX3BF7P9OPtrjpo7vE8FPbKDLUNanD5PL72QcpUvT1Pn4Mku+fhutH9W+1b690F0oDn7zLXPMi36k2yexyBfKreoNavhi1deJ7qx8ikNvD+oHtAdfG2Ep1MMz14Hl15K7qX7KT3+Apx0kmVU+0n0I3kHqWr5X3//ff/et/pp6O2retL1L5S5YVaJ+lli6pjjQtR5fMgfQpkJ314OPyHPzeNFUIId+Con+N/PVoQwusKj/P44z0hHCV9KYSHTxycEMIQnr3uPugZwusEz9DzbPyez9KE0JG+FMLrQb5kIITQwi7lfX47UAjXgA+i86HmGKThXNKXQnh9yAlOCCGEEEII4dFwVQeH7wwPIYQQQgghPG4e0plJTnBCCCGEEEIIj4azHZyc0oQQQgghhBBG3Pd5Sk5wQgghhBBCCI+GfItaCCGEEEII4dEQByeEEEIIIYTwaIiDE0IIIYQQQngkPHnyX8G/AwQjWCuWAAAAAElFTkSuQmCC"

    for record in oio_records:
        subject = 'OIO to be issued as ph2 time is over'
        message = f"please send OIO for  bond number {record.bond_number} ."
        from_email = settings.EMAIL_HOST_USER  # Use the configured from_email from settings
        recipient_list = [settings.office_mail_id]

        html_content = render_to_string('oio.html',
                                        {'data': record, 'base64_image': base64_image, 'DC': settings.deputy_cmsnr,
                                         "present_date": datetime.now(timezone("Asia/Kolkata")).strftime('%d-%m-%Y')})

        # Generate the PDF from the HTML content
        # pdf_stream = generate_pdf(html_content)
        # pdf_bytes = pdf_stream.getvalue()

        # Generate the DOCX file from HTML
        docx_byte_data = generate_docx_from_html(html_content)

        # if pdf_bytes:
        #     # Attach the PDF to the email
        #     pdf_file = BytesIO(pdf_bytes)
        #     pdf_filename = f'OIO_{record.bond_number}.pdf'
        if docx_byte_data:
            email = EmailMessage(subject, message, from_email, recipient_list)
            # email.attach(pdf_filename, pdf_file.read(), 'application/pdf')
            email.attach('oio_epcg_reminder.docx', docx_byte_data,
                         'application/vnd.openxmlformats-officedocument.wordprocessingml.document')

            # Send the email
            email.send()
        record.is_oio_issued = True
        record.save()

    for record in decc_oio_records:
        subject = 'OIO to be issued as ph2 time is over'
        message = f"please send OIO for  bond number {record.bond_number} ."
        from_email = settings.EMAIL_HOST_USER  # Use the configured from_email from settings
        recipient_list = [settings.office_mail_id]  # [record.gmail_id]

        html_content = render_to_string('oio.html',
                                        {'data': record, 'base64_image': base64_image, 'DC': settings.deputy_cmsnr,
                                         "present_date": datetime.now(timezone("Asia/Kolkata")).strftime('%d-%m-%Y')})

        # Generate the PDF from the HTML content
        # pdf_stream = generate_pdf(html_content)
        # pdf_bytes = pdf_stream.getvalue()
        # Generate the DOCX file from HTML
        docx_byte_data = generate_docx_from_html(html_content)

        # if pdf_bytes:
        #     # Attach the PDF to the email
        #     pdf_file = BytesIO(pdf_bytes)
        #     pdf_filename = f'OIO_{record.bond_number}.pdf'
        if docx_byte_data:
            email = EmailMessage(subject, message, from_email, recipient_list)
            # email.attach(pdf_filename, pdf_file.read(), 'application/pdf')
            email.attach('oio_decc_reminder.docx', docx_byte_data,
                         'application/vnd.openxmlformats-officedocument.wordprocessingml.document')

            # Send the email
            email.send()
        record.is_oio_issued = True
        record.save()

    for record in ph2_records:
        subject = 'PH2 to be issued as PH1 time is over'
        message = f"please send PH2 for  bond number {record.bond_number} ."
        from_email = settings.EMAIL_HOST_USER  # Use the configured from_email from settings
        recipient_list = [settings.office_mail_id]  # [record.gmail_id]

        html_content = render_to_string('ph2.html',
                                        {'data': record, 'base64_image': base64_image, 'DC': settings.deputy_cmsnr,
                                         "present_date": datetime.now(timezone("Asia/Kolkata")).strftime('%d-%m-%Y')})

        # Generate the PDF from the HTML content
        # pdf_stream = generate_pdf(html_content)
        # pdf_bytes = pdf_stream.getvalue()
        # Generate the DOCX file from HTML
        docx_byte_data = generate_docx_from_html(html_content)

        # if pdf_bytes:
        #     # Attach the PDF to the email
        #     pdf_file = BytesIO(pdf_bytes)
        #     pdf_filename = f'PH2_{record.bond_number}.pdf'

        # Create the email message with attached PDF
        if docx_byte_data:
            email = EmailMessage(subject, message, from_email, recipient_list)
            # email.attach(pdf_filename, pdf_file.read(), 'application/pdf')
            email.attach('ph2_epcg_reminder.docx', docx_byte_data,
                         'application/vnd.openxmlformats-officedocument.wordprocessingml.document')

            # Send the email
            email.send()
        record.is_ph2_issued = True
        record.save()

    for record in decc_ph2_records:
        subject = 'PH2 to be issued as PH1 time is over'
        message = f"please send PH2 for  bond number {record.bond_number} ."
        from_email = settings.EMAIL_HOST_USER  # Use the configured from_email from settings
        recipient_list = [settings.office_mail_id]  # [record.gmail_id]

        html_content = render_to_string('ph2.html',
                                        {'data': record, 'base64_image': base64_image, 'DC': settings.deputy_cmsnr,
                                         "present_date": datetime.now(timezone("Asia/Kolkata")).strftime('%d-%m-%Y')})

        # Generate the PDF from the HTML content
        # pdf_stream = generate_pdf(html_content)
        # pdf_bytes = pdf_stream.getvalue()
        # Generate the DOCX file from HTML
        docx_byte_data = generate_docx_from_html(html_content)

        # if pdf_bytes:
        #     # Attach the PDF to the email
        #     pdf_file = BytesIO(pdf_bytes)
        #     pdf_filename = f'PH2_{record.bond_number}.pdf'
        # Create the email message with attached PDF
        if docx_byte_data:
            email = EmailMessage(subject, message, from_email, recipient_list)
            # email.attach(pdf_filename, pdf_file.read(), 'application/pdf')
            email.attach('ph2_decc_reminder.docx', docx_byte_data,
                         'application/vnd.openxmlformats-officedocument.wordprocessingml.document')

            # Send the email
            email.send()
        record.is_ph2_issued = True
        record.save()

    for record in ph1_records:
        subject = 'PH1 to be issued as SCN time is over'
        message = f"please send PH1 for  bond number {record.bond_number} ."
        from_email = settings.EMAIL_HOST_USER  # Use the configured from_email from settings
        recipient_list = [settings.office_mail_id]

        html_content = render_to_string('ph1.html',
                                        {'data': record, 'base64_image': base64_image, 'DC': settings.deputy_cmsnr,
                                         "present_date": datetime.now(timezone("Asia/Kolkata")).strftime('%d-%m-%Y')})

        # Generate the PDF from the HTML content
        # pdf_stream = generate_pdf(html_content)
        # pdf_bytes = pdf_stream.getvalue()
        # Generate the DOCX file from HTML
        docx_byte_data = generate_docx_from_html(html_content)

        # if pdf_bytes:
        #     # Attach the PDF to the email
        #     pdf_file = BytesIO(pdf_bytes)
        #     pdf_filename = f'PH1_{record.bond_number}.pdf'
        # Create the email message with attached PDF
        if docx_byte_data:
            email = EmailMessage(subject, message, from_email, recipient_list)
            # email.attach(pdf_filename, pdf_file.read(), 'application/pdf')
            email.attach('ph1_epcg_reminder.docx', docx_byte_data,
                         'application/vnd.openxmlformats-officedocument.wordprocessingml.document')

            # Send the email
            email.send()
        record.is_ph1_issued = True
        record.save()

    for record in decc_ph1_records:
        subject = 'PH1 to be issued as SCN time is over'
        message = f"please send PH1 for  bond number {record.bond_number} ."
        from_email = settings.EMAIL_HOST_USER  # Use the configured from_email from settings
        recipient_list = [settings.office_mail_id]

        html_content = render_to_string('ph1.html',
                                        {'data': record, 'base64_image': base64_image, 'DC': settings.deputy_cmsnr,
                                         "present_date": datetime.now(timezone("Asia/Kolkata")).strftime('%d-%m-%Y')})

        # Generate the PDF from the HTML content
        # pdf_stream = generate_pdf(html_content)
        # pdf_bytes = pdf_stream.getvalue()
        # Generate the DOCX file from HTML
        docx_byte_data = generate_docx_from_html(html_content)

        # if pdf_bytes:
        #     # Attach the PDF to the email
        #     pdf_file = BytesIO(pdf_bytes)
        #     pdf_filename = f'PH1_{record.bond_number}.pdf'
        # Create the email message with attached PDF
        if docx_byte_data:
            email = EmailMessage(subject, message, from_email, recipient_list)
            # email.attach(pdf_filename, pdf_file.read(), 'application/pdf')
            email.attach('ph1_decc_reminder.docx', docx_byte_data,
                         'application/vnd.openxmlformats-officedocument.wordprocessingml.document')

            # Send the email
            email.send()
        record.is_ph1_issued = True
        record.save()

    for record in scn_records:
        subject = 'SCN to be issued as license is expired'
        message = f"please send SCN for  bond number {record.bond_number} ."
        from_email = settings.EMAIL_HOST_USER  # Use the configured from_email from settings
        recipient_list = [settings.office_mail_id]

        html_content = render_to_string('SCN.html',
                                        {'data': record, 'base64_image': base64_image, 'DC': settings.deputy_cmsnr,
                                         "present_date": datetime.now(timezone("Asia/Kolkata")).strftime('%d-%m-%Y')})

        # Generate the PDF from the HTML content
        # pdf_stream = generate_pdf(html_content)
        # pdf_bytes = pdf_stream.getvalue()
        # Generate the DOCX file from HTML
        docx_byte_data = generate_docx_from_html(html_content)

        # if pdf_bytes:
        #     # Attach the PDF to the email
        #     pdf_file = BytesIO(pdf_bytes)
        #     pdf_filename = f'SCN_{record.bond_number}.pdf'
        # Create the email message with attached PDF
        if docx_byte_data:
            email = EmailMessage(subject, message, from_email, recipient_list)
            # email.attach(pdf_filename, pdf_file.read(), 'application/pdf')
            email.attach('scn_epcg_reminder.docx', docx_byte_data,
                         'application/vnd.openxmlformats-officedocument.wordprocessingml.document')

            # Send the email
            email.send()
        record.is_scn_issued = True
        record.save()

    for record in decc_scn_records:
        subject = 'SCN to be issued as license is expired'
        message = f"please send SCN for  bond number {record.bond_number} ."
        from_email = settings.EMAIL_HOST_USER  # Use the configured from_email from settings
        recipient_list = [settings.office_mail_id]

        html_content = render_to_string('SCN.html',
                                        {'data': record, 'base64_image': base64_image, 'DC': settings.deputy_cmsnr,
                                         "present_date": datetime.now(timezone("Asia/Kolkata")).strftime('%d-%m-%Y')})

        # Generate the PDF from the HTML content
        # pdf_stream = generate_pdf(html_content)
        # pdf_bytes = pdf_stream.getvalue()

        # Generate the DOCX file from HTML
        docx_byte_data = generate_docx_from_html(html_content)

        # if pdf_bytes:
        #     # Attach the PDF to the email
        #     pdf_file = BytesIO(pdf_bytes)
        #     pdf_filename = f'SCN_{record.bond_number}.pdf'
        # Create the email message with attached PDF
        if docx_byte_data:
            email = EmailMessage(subject, message, from_email, recipient_list)
            # email.attach(pdf_filename, pdf_file.read(), 'application/pdf')
            # Attach the DOCX file
            email.attach('scn_decc_reminder.docx', docx_byte_data,
                         'application/vnd.openxmlformats-officedocument.wordprocessingml.document')

            # Send the email
            email.send()
        record.is_scn_issued = True
        record.save()

    for record in letter_records:
        subject = 'License Expiry Reminder'
        message = f"Your license for bond number {record.bond_number} is about to expire."
        from_email = settings.EMAIL_HOST_USER  # Use the configured from_email from settings
        recipient_list = [record.gmail_id]

        html_content = render_to_string('letter.html',
                                        {'data': record, 'base64_image': base64_image, 'DC': settings.deputy_cmsnr,
                                         "present_date": datetime.now(timezone("Asia/Kolkata").strftime('%d-%m-%Y'))})

        # Generate the PDF from the HTML content
        pdf_stream = generate_pdf(html_content)
        pdf_bytes = pdf_stream.getvalue()

        if pdf_bytes:
            # Attach the PDF to the email
            pdf_file = BytesIO(pdf_bytes)
            pdf_filename = f'license_reminder_{record.bond_number}.pdf'

            # Create the email message with attached PDF
            email = EmailMessage(subject, message, from_email, recipient_list)
            email.attach(pdf_filename, pdf_file.read(), 'application/pdf')

            # Send the email
            email.send()
        record.is_letter_issued = True
        record.save()

    for record in decc_letter_records:
        subject = 'License Expiry Reminder'
        message = f"Your license for bond number {record.bond_number} is about to expire."
        from_email = settings.EMAIL_HOST_USER  # Use the configured from_email from settings
        recipient_list = [record.gmail_id]

        html_content = render_to_string('letter.html',
                                        {'data': record, 'base64_image': base64_image, 'DC': settings.deputy_cmsnr,
                                         "present_date": datetime.now(timezone("Asia/Kolkata")).strftime('%d-%m-%Y')})

        # Generate the PDF from the HTML content
        pdf_stream = generate_pdf(html_content)
        pdf_bytes = pdf_stream.getvalue()

        if pdf_bytes:
            # Attach the PDF to the email
            pdf_file = BytesIO(pdf_bytes)
            pdf_filename = f'license_reminder_{record.bond_number}.pdf'

            # Create the email message with attached PDF
            email = EmailMessage(subject, message, from_email, recipient_list)
            email.attach(pdf_filename, pdf_file.read(), 'application/pdf')

            # Send the email
            email.send()
        record.is_letter_issued = True
        record.save()
    # Redirect back to the home page
    return redirect('home')


id = -1
logged_in = 0


def get_loogged_in():
    global logged_in
    return logged_in


def set_logged_in(status):
    global logged_in
    logged_in = status


def get_id():
    return id


def set_id(id_no):
    global id
    id = id_no


@never_cache
def custom_login_view(request):
    global logged_in
    # logged_in = 0
    set_logged_in(0)

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            set_logged_in(1)
            return redirect('home')
    return render(request, 'login.html')


@never_cache
@login_required
def home(request):
    if get_loogged_in() == 0:
        return redirect('login')
    form = ImporterForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        form.save()

    form = ImporterForm()
    context = {'form': form}
    return render(request, 'home.html', context)


@never_cache
@login_required
def search_modify(request):
    if get_loogged_in() == 0:
        return redirect('login')
    form = ImporterForm(request.POST or None)
    importer = None
    if request.method == 'POST' and form.is_valid():
        bond_id = form.cleaned_data['bond_id']
        importer = get_object_or_404(Importer, bond_number=bond_id)
        form = ImporterForm(instance=importer)

    context = {'form': form, 'importer': importer}
    return render(request, 'search_modify.html', context)


@never_cache
@login_required
def modify_data(request):
    if get_loogged_in() == 0:
        return redirect('login')
    data = {}
    if request.method == 'POST':
        form = Importer.objects.filter(bond_number=request.POST['bond_id']).order_by('-lic_date')

        for i in form:
            set_id(i.id)

            data['importer_name'] = i.importer_name
            data['address'] = i.address
            data['license_number'] = i.license_number
            data['lic_date'] = i.lic_date
            data['bond_number'] = i.bond_number
            data['gmail_id'] = i.gmail_id
            data['radio_choice'] = i.radio_choice
            data['is_eodc_produced'] = i.is_eodc_produced
            data['is_dgft_ack_produced'] = i.is_dgft_ack_produced
            data['is_letter_issued'] = i.is_letter_issued
            data['is_scn_issued'] = i.is_scn_issued
            data['is_ph1_issued'] = i.is_ph1_issued
            data['is_ph2_issued'] = i.is_ph2_issued
            data['is_oio_issued'] = i.is_oio_issued
            data['is_paused'] = i.is_paused
            data['is_closed'] = i.is_closed
            break

        context = {'present': data}

        if not bool(data):
            return render(request, 'search_modify.html')
        return render(request, 'modify_data.html', context)
    else:
        return render(request, 'search_modify.html')


@never_cache
@login_required
def update_data(request):
    if get_loogged_in() == 0:
        return redirect('login')
    if request.method == 'POST':
        form = Importer.objects.get(id=id)

        if bool(request.POST.get('is_eodc_produced')):
            if request.POST.get('is_eodc_produced') == 'True':
                form.is_eodc_produced = True
            else:
                form.is_eodc_produced = False

        if bool(request.POST.get('is_dgft_ack_produced')):
            if request.POST.get('is_dgft_ack_produced') == 'True':
                form.is_dgft_ack_produced = True
            else:
                form.is_dgft_ack_produced = False

        if bool(request.POST.get('is_letter_issued')):
            if request.POST.get('is_letter_issued') == 'True':
                form.is_letter_issued = True
            else:
                form.is_letter_issued = False

        if bool(request.POST.get('is_scn_issued')):
            if request.POST.get('is_scn_issued') == 'True':
                form.is_scn_issued = True
            else:
                form.is_scn_issued = False

        if bool(request.POST.get('is_ph1_issued')):
            if request.POST.get('is_ph1_issued') == 'True':
                form.is_ph1_issued = True
            else:
                form.is_ph1_issued = False

        if bool(request.POST.get('is_ph2_issued')):
            if request.POST.get('is_ph2_issued') == 'True':
                form.is_ph2_issued = True
            else:
                form.is_ph2_issued = False

        if bool(request.POST.get('is_oio_issued')):
            if request.POST.get('is_oio_issued') == 'True':
                form.is_oio_issued = True
            else:
                form.is_oio_issued = False

        if bool(request.POST.get('is_paused')):
            if request.POST.get('is_paused') == 'True':
                form.is_paused = True
            else:
                form.is_paused = False

        if bool(request.POST.get('is_closed')):
            if request.POST.get('is_closed') == 'True':
                form.is_closed = True
            else:
                form.is_closed = False

        form.save()
        context = {'present': form}
        return render(request, 'modify_data.html', context)
    else:
        return render(request, 'search_modify.html')
